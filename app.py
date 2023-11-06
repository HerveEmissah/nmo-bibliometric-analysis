from flask import Flask, jsonify, request, json, url_for, Response, render_template, abort, make_response
from flask_pymongo import PyMongo
from flask_mail import Mail, Message
import pymongo
from pymongo import MongoClient
import os
#from bson.objectid import ObjectId
from datetime import datetime
import urllib.request, json
from urllib.request import urlopen
import socket
import xml.etree.ElementTree as ET
import asyncio
import time
import collections
import mechanize
import json
from bson import json_util
import math
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import numbers
from pycel import ExcelCompiler
import urllib.parse
from urllib.parse import urlencode
import datetime
from itsdangerous import URLSafeSerializer, BadData
import re
import requests
import subprocess
import sshtunnel
from sshtunnel import SSHTunnelForwarder

app = Flask(__name__)

app.config["MONGO_URI"] = "mongodb://localhost:27017/NMODatabase"
client = MongoClient(os.environ['DB_PORT_27017_TCP_ADDR'], 27017)
db = client.NMODatabase

mongo = PyMongo(app)

app.config.update(
    DEBUG=True,
    #EMAIL SETTINGS
    MAIL_SERVER='smtp.gmail.com',
    MAIL_PORT=465,
    MAIL_USE_TLS=False,
    MAIL_USE_SSL=True,
    MAIL_USERNAME = 'NMOBiblio@gmail.com',
    #MAIL_PASSWORD = 'NMOpassword'
    MAIL_PASSWORD = 'nbzccvztaujrbgei'  #Gmail 2-Step Authentication
)

mail = Mail(app)

# Configuration for the MySQL database
db_config = {
    'host': 'LoggingData',
    'user': 'blxps',
    'password': '100%db',
    'db_name': 'LoggingData'
}

# Configuration for the SSH tunnel to the remote server
ssh_config = {
    'ssh_address': ('129.174.130.34', 22),
    'ssh_username': 'herve1',
    'ssh_password': 'herve123',
    'remote_bind_address': ('localhost', 3306),  # Assuming MySQL uses default port
}

dump_file_path = '/nmo/tables_dump'


#create directory with write permissions (in octal notation, 0o755 allows owner to read/write/execute, others to read/execute)
if not os.path.exists(dump_file_path):
    os.makedirs(dump_file_path, mode=0o755)


@app.route('/users/neuromorphodata', methods=["POST"])
def neuromorphodata():

    citing = db.NMO_Citing
    using = db.NMO_Using
    describing = db.NMO_Describing

    #Fetch literature Citing Neuromorpho
    dataCitingPageUrl = "http://cng.gmu.edu:8080/apiLiteratureMain/literature/articles?data.dataUsage=CITING&page=0"
    urlHandlerCiting = urllib.request.urlopen(dataCitingPageUrl)
    if urlHandlerCiting.getcode()==200:
        data = urlHandlerCiting.read()
        jsonData = json.loads(data)

        #iterate thru each page
        totPagesCiting = jsonData["totalPages"]
        CitingRecCount = 0
        citing.drop()
        for i in range (0, totPagesCiting):
           dataCitingUrl = "http://cng.gmu.edu:8080/apiLiteratureMain/literature/articles?data.dataUsage=CITING&page=" + str(i)
           urlHandlerCiting = urllib.request.urlopen(dataCitingUrl)
           if urlHandlerCiting.getcode()==200:
              data = urlHandlerCiting.read()
              jsonData = json.loads(data)
              for i in jsonData["content"]:
                 citing.insert_one({
                    'doi': i["doi"],
                    'pmid': i["pmid"],
                    'title': i["title"],
                    'publishedDate': i["publishedDate"]
                 })
                 CitingRecCount +=1
           else:
              result = jsonify({"error":"Error receiving data: " + urlHandlerCiting.getcode()})

    else:
        result = jsonify({"error":"Error receiving data: " + urlHandlerCiting.getcode()})

    urlHandlerCiting.close()

    #Fetch literature Using Neuromorpho
    dataUsingPageUrl = "http://cng.gmu.edu:8080/apiLiteratureMain/literature/articles?data.dataUsage=USING&page=0"
    urlHandlerUsing = urllib.request.urlopen(dataUsingPageUrl)
    if urlHandlerUsing.getcode()==200:
        data = urlHandlerUsing.read()
        jsonData = json.loads(data)

        #iterate thru each page
        totPagesUsing = jsonData["totalPages"]
        UsingRecCount = 0
        using.drop()
        for i in range (0, totPagesUsing):
           dataUsingUrl = "http://cng.gmu.edu:8080/apiLiteratureMain/literature/articles?data.dataUsage=USING&page=" + str(i)
           urlHandlerUsing = urllib.request.urlopen(dataUsingUrl)
           if urlHandlerUsing.getcode()==200:
              data = urlHandlerUsing.read()
              jsonData = json.loads(data)
              for i in jsonData["content"]:
                 using.insert_one({
                    'doi': i["doi"],
                    'pmid': i["pmid"],
                    'title': i["title"],
                    'publishedDate': i["publishedDate"]
                 })
                 UsingRecCount +=1
           else:
              result = jsonify({"error":"Error receiving data: " + urlHandlerUsing.getcode()})

    else:
        result = jsonify({"error":"Error receiving data: " + urlHandlerUsing.getcode()})

    urlHandlerUsing.close()

    #Fetch literature Describing Neuromorpho
    dataDescribingPageUrl = "http://129.174.10.65:8189/articles/status/Positive?reconstructions.currentStatusList.specificDetails=In%20repository&page=0"
    urlHandlerDescribing = urllib.request.urlopen(dataDescribingPageUrl)
    if urlHandlerDescribing.getcode()==200:
        data = urlHandlerDescribing.read()
        jsonData = json.loads(data)

        #iterate thru each page
        totPagesDescribing = jsonData["totalPages"]
        DescRecCount = 0
        describing.drop()
        for i in range (0, totPagesDescribing):
           dataDescribingUrl = "http://129.174.10.65:8189/articles/status/Positive?reconstructions.currentStatusList.specificDetails=In%20repository&page=" + str(i)
           urlHandlerDescribing = urllib.request.urlopen(dataDescribingUrl)
           if urlHandlerDescribing.getcode()==200:
              data = urlHandlerDescribing.read()
              jsonData = json.loads(data)

              for content in jsonData["content"]:
                 for key,value in content["data"].items():
                     if key == 'pmid': pmid = value
                     else:
                         if key == 'doi' : doi = value
                         else:
                             if key == 'title' : title = value
                             else:
                                 if key == 'publishedDate' : publishedDate = value

                 describing.insert_one({
                        'doi': doi,
                        'pmid': pmid,
                        'title': title,
                        'publishedDate': publishedDate
                 })
                 DescRecCount +=1
           else:
              result = jsonify({"error":"Error receiving data: " + urlHandlerDescribing.getcode()})

    else:
        result = jsonify({"error":"Error receiving data: " + urlHandlerDescribing.getcode()})

    urlHandlerDescribing.close()

    #update the null pmid if doi is known
    doi_citing=None
    pmid_citing=None
    for doc in citing.find():
        if doc["doi"] != None:
            if doc["pmid"] == None:
                doi_citing = doc["doi"].strip("\u00a0") #remove any potential leading or trailing blanks encoding
                pmid_citing = get_pmid(doi_citing)
                citing.update_one({"_id": doc["_id"]}, { "$set": {"pmid": pmid_citing} })

    doi_using=None
    pmid_using=None
    for doc in citing.find():
        if doc["doi"] != None:
            if doc["pmid"] == None:
                doi_using = doc["doi"].strip("\u00a0")
                pmid_using = get_pmid(doi_using)
                citing.update_one({"_id": doc["_id"]}, { "$set": {"pmid": pmid_using} })

    doi_describing=None
    pmid_describing=None
    for doc in citing.find():
        if doc["doi"] != None:
            if doc["pmid"] == None:
                doi_describing = doc["doi"].strip("\u00a0")
                pmid_describing = get_pmid(doi_describing)
                citing.update_one({"_id": doc["_id"]}, { "$set": {"pmid": pmid_describing} })

    #update the null doi if pmid is known
    doi_citing=None
    pmid_citing=None
    for doc in citing.find():
        if doc["pmid"] != None:
            if doc["doi"] == None:
                pmid_citing = doc["pmid"].strip("\u00a0") #remove any potential leading or trailing blanks encoding
                doi_citing = get_doi(pmid_citing)
                citing.update_one({"_id": doc["_id"]}, { "$set": {"doi": doi_citing} })

    doi_using=None
    pmid_using=None
    for doc in citing.find():
        if doc["pmid"] != None:
            if doc["doi"] == None:
                pmid_using = doc["pmid"].strip("\u00a0")
                doi_using = get_doi(pmid_using)
                citing.update_one({"_id": doc["_id"]}, { "$set": {"doi": doi_using} })

    doi_describing=None
    pmid_describing=None
    for doc in citing.find():
        if doc["pmid"] != None:
            if doc["doi"] == None:
                pmid_describing = doc["pmid"].strip("\u00a0")
                doi_describing = get_doi(pmid_describing)
                citing.update_one({"_id": doc["_id"]}, { "$set": {"doi": doi_describing} })



    result = "records inserted: " + str(CitingRecCount) + "  in DB.Citing, " + str(UsingRecCount) + " in DB.Using and " + str(DescRecCount) + " in DB.Describing"

    return result


def getDescribingOverallCitedBy(doi):

    crossref_Url = 'http://api.crossref.org/works/' + str(doi)
    br = mechanize.Browser()
    br.set_handle_robots(False) # ignore robots
    br.addheaders = [('User-agent', 'Mozilla/5.0 (X11; U; Linux i686; en-US; rv:1.9.0.1) Gecko/2008071615 Fedora/3.0.1-1.fc9 Firefox/3.0.1')]

    overallReferencedBy = None
    try:
        br.open(crossref_Url)
        res = br.response()
        jsonData = json.loads(res.read())

        for key, value in jsonData["message"].items():
            #print(key,":",value)
            if key == 'is-referenced-by-count':
                overallReferencedBy = value
                print ('overallReferencedBy:', overallReferencedBy)

    except (mechanize.HTTPError,mechanize.URLError) as e:
      if isinstance(e,mechanize.HTTPError):
          print ('ErrorCode:', e.code,' url: ', crossref_Url)
      else:
          print ('ErrorReason:', e.reason.args,' url: ', crossref_Url)

    return overallReferencedBy


def get_doi(pmid):

    doi = None
    europePMC_Url = 'https://www.ebi.ac.uk/europepmc/webservices/rest/search?query=' + str(pmid) + '&format=json'
    urlHandler = urllib.request.urlopen(europePMC_Url)
    if urlHandler.getcode()==200:
        data = urlHandler.read()
        jsonData = json.loads(data)
        resultList = jsonData['resultList']['result']
        for elem in resultList:
            try:
                if elem['id'] == pmid:
                    doi = elem['doi']
                    break
            except KeyError: pass

    urlHandler.close()

    return doi


def get_pubDate(pmid):

    pubDate = None
    europePMC_Url = 'https://www.ebi.ac.uk/europepmc/webservices/rest/search?query=' + str(pmid) + '&format=json'
    urlHandler = urllib.request.urlopen(europePMC_Url)
    if urlHandler.getcode()==200:
        data = urlHandler.read()
        jsonData = json.loads(data)
        resultList = jsonData['resultList']['result']
        for elem in resultList:
            try:
                if elem['id'] == pmid:
                    pubDate = elem['firstPublicationDate']
                    break
            except KeyError: pass

    urlHandler.close()

    return pubDate


def get_pmid(doi):

    pmid = None
    europePMC_Url = 'https://www.ebi.ac.uk/europepmc/webservices/rest/search?query=' + str(doi) + '&format=json'
    #print(europePMC_Url)
    urlHandler = urllib.request.urlopen(europePMC_Url)
    if urlHandler.getcode()==200:
        data = urlHandler.read()
        jsonData = json.loads(data)
        resultList = jsonData['resultList']['result']
        for elem in resultList:
            try:
                if elem['id'] == doi:
                    pmid = elem['pmid']
                    break
            except KeyError: pass

    urlHandler.close()

    return pmid


async def fetchReferences_crossref(url, NMO_Type):
  #await asyncio.sleep(1)

  citing = db.NMO_Citing
  NMOCiting_References = db.NMOCiting_References

  using = db.NMO_Using
  NMOUsing_References = db.NMOUsing_References

  br = mechanize.Browser()
  br.set_handle_robots(False) # ignore robots
  br.addheaders = [('User-agent', 'Mozilla/5.0 (X11; U; Linux i686; en-US; rv:1.9.0.1) Gecko/2008071615 Fedora/3.0.1-1.fc9 Firefox/3.0.1')]

  try:
      br.open(url)
      res = br.response()
      jsonData = json.loads(res.read())

      child_pmid = None
      child_doi = None
      child_pubYear = None
      referencesObj = []
      references = []
      for key, value in jsonData["message"].items():
          #print(key,":",value)
          if key == 'reference':
              for i in range (0, len(value)):
                  try:
                      child_doi = value[i]["DOI"]
                      child_pubYear = value[i]["year"]
                  except KeyError: pass

                  #child_pmid = get_pmid(child_doi)
                  referencesObj = {'pmid': child_pmid, 'doi': child_doi, 'pubYear': child_pubYear}
                  references.append(referencesObj)

      doi_start = url.find("works/")
      doi_end = len(url)
      doi = url[doi_start+6 : doi_end]
      pmid = None

      if NMO_Type == "NMO_Citing":
          new_rec_id = NMOCiting_References.insert_one({
                            'pmid': pmid,
                            'doi' : doi,
                            'references': references,
                            'title': None,
                            'publishedDate': None,
                            'source': 'crossref'
                            })

          #update metadata: pmid, title, publishedDate
          for doc in citing.find():
              if doc["doi"] == doi:
                  NMOCiting_References.update_one({"_id": new_rec_id.inserted_id}, { "$set": {
                                                                                     "pmid": doc["pmid"] ,
                                                                                     "title": doc["title"],
                                                                                     "publishedDate": doc["publishedDate"] } })

      elif NMO_Type == "NMO_Using":
          new_rec_id = NMOUsing_References.insert_one({
                            'pmid': pmid,
                            'doi' : doi,
                            'references': references,
                            'title': None,
                            'publishedDate': None,
                            'source': 'crossref'
                            })

          #update metadata: pmid, title, publishedDate
          for doc in using.find():
              if doc["doi"] == doi:
                  NMOUsing_References.update_one({"_id": new_rec_id.inserted_id}, { "$set": {
                                                                                    "pmid": doc["pmid"] ,
                                                                                    "title": doc["title"],
                                                                                    "publishedDate": doc["publishedDate"] } })

  except (mechanize.HTTPError,mechanize.URLError) as e:
      if isinstance(e,mechanize.HTTPError):
          print ('ErrorCode:', e.code,' url: ', url)
      else:
          print ('ErrorReason:', e.reason.args,' url: ', url)


async def fetchReferences_europepmc(url, NMO_Type):
  #await asyncio.sleep(1)

  citing = db.NMO_Citing
  NMOCiting_References = db.NMOCiting_References

  using = db.NMO_Using
  NMOUsing_References = db.NMOUsing_References

  print ('url:', url)
  var_url = urlopen(url)
  xmldoc = ET.parse(var_url)
  pmidstart = url.find("MED/")
  pmidend = url.find("/references")  #references in europepmc are the listed citations of the given literature
  pmid = url[pmidstart+4 : pmidend]
  doi = None
  #child_doi = None

  references = []
  root = xmldoc.getroot()
  hitCount = root.find('hitCount').text
  req = root.find('request')
  pageSize = req.find('pageSize').text
  if int(hitCount) > 0:
      #find number of pages based on hitCount and pageSize
      if (int(hitCount) % int(pageSize) == 0):
          numPage = int(int(hitCount) / int(pageSize))
      else:
          numPage = int(math.floor(int(hitCount) / int(pageSize)) + 1)
      referenceList = root.find('referenceList')
      for elems in referenceList:
          child_pmid = None
          for Ids in elems:
              if Ids.tag == 'id':
                  child_pmid = Ids.text
                  #child_doi = get_doi(child_pmid)
              if Ids.tag == 'pubYear':
                  child_pubYear = Ids.text
          referencesObj = {'pmid': child_pmid, 'pubYear': child_pubYear}
          references.append(referencesObj)

      #Get next page if pageSize > 1
      for i in range (1, numPage):
          urlNext = url + '?page=' + str(i+1)
          print('url:' , urlNext)
          var_urlNext = urlopen(urlNext)
          xmldocNext = ET.parse(var_urlNext)
          rootNext = xmldocNext.getroot()
          referenceListNext = rootNext.find('referenceList')
          for elems in referenceListNext:
              child_pmid = None
              for Ids in elems:
                  if Ids.tag == 'id':
                      child_pmid = Ids.text
                      #child_doi = get_doi(child_pmid)
                      referencesObj = {'pmid': child_pmid, 'pubYear': child_pubYear}
                      references.append(referencesObj)

      if NMO_Type == "NMO_Citing":
          new_rec_id = NMOCiting_References.insert_one({
                            'pmid': pmid,
                            'doi' : doi,
                            'references': references,
                            'title': None,
                            'publishedDate': None,
                            'source': 'europepmc'
                            })

          #update metadata: doi, title, publishedDate
          for doc in citing.find():
              if doc["pmid"] == pmid:
                  NMOCiting_References.update_one({"_id": new_rec_id.inserted_id}, { "$set": {
                                                                                     "doi": doc["doi"] ,
                                                                                     "title": doc["title"],
                                                                                     "publishedDate": doc["publishedDate"] } })

      elif NMO_Type == "NMO_Using":
          new_rec_id = NMOUsing_References.insert_one({
                            'pmid': pmid,
                            'doi' : doi,
                            'references': references,
                            'title': None,
                            'publishedDate': None,
                            'source': 'europepmc'
                            })

          #update metadata: pmid, title, publishedDate
          for doc in using.find():
              if doc["pmid"] == pmid:
                  NMOUsing_References.update_one({"_id": new_rec_id.inserted_id}, { "$set": {
                                                                                    "doi": doc["doi"] ,
                                                                                    "title": doc["title"],
                                                                                    "publishedDate": doc["publishedDate"] } })


async def fetchCitedBy_europepmc(url):
  await asyncio.sleep(1)

  #timeout in seconds
  timeout = 30
  socket.setdefaulttimeout(timeout)

  describing = db.NMO_Describing
  NMODescribing_Overall_CitedBy = db.NMODescribing_Overall_CitedBy

  print ('url:', url)
  hitCount = 0
  pageSize = 0
  try:
     var_url = urlopen(url)
     xmldoc = ET.parse(var_url)

     pmidstart = url.find("MED/")
     pmidend = url.find("/citations")
     pmid = url[pmidstart+4 : pmidend]
     print ('pmid:', pmid)
     doi = None

     root = xmldoc.getroot()
     hitCount = int(root.find('hitCount').text)
     req = root.find('request')
     pageSize = int(req.find('pageSize').text)
     print ('hitCount:', hitCount)
     print ('pageSize:', pageSize)
  except urllib.error.URLError as err:
     print("error in url", url)
     print ('Err url:', url)

  if hitCount > 0:
      citedBy = []
      #find number of pages based on hitCount and pageSize
      if hitCount % pageSize == 0:
          numPage = int(hitCount / pageSize)
      else:
          numPage = int(math.floor(hitCount / pageSize) + 1)
      citedByList = root.find('citationList')
      for elems in citedByList:
          citedByObj = {}
          child_pmid = None
          child_puYear = None
          child_title = None
          child_authorString = None
          child_journalAbbreviation = None
          for Ids in elems:
              if Ids.tag == 'id':
                  child_pmid = Ids.text
              if Ids.tag == 'pubYear':
                  child_pubYear = Ids.text
              if Ids.tag == 'title':
                  child_title = Ids.text
              if Ids.tag == 'authorString':
                  child_authorString = Ids.text
              if Ids.tag == 'journalAbbreviation':
                  child_journalAbbreviation = Ids.text

          citedByObj = {'pmid': child_pmid, 'pubYear': child_pubYear, 'title': child_title, 'authorString': child_authorString, 'journalAbbreviation': child_journalAbbreviation}
          citedBy.append(citedByObj)

      #Get next page if pageSize > 1
      for i in range (1, numPage):
          await asyncio.sleep(3)
          urlNext = url + '?page=' + str(i+1)
          #print('url:' , urlNext)
          var_urlNext = urlopen(urlNext, timeout=30)
          xmldocNext = ET.parse(var_urlNext)
          rootNext = xmldocNext.getroot()
          citedByListNext = rootNext.find('citationList')
          for elems in citedByListNext:
              citedByObj = {}
              child_pmid = None
              child_puYear = None
              child_title = None
              for Ids in elems:
                  if Ids.tag == 'id':
                      child_pmid = Ids.text
                  if Ids.tag == 'pubYear':
                      child_pubYear = Ids.text
                  if Ids.tag == 'title':
                      child_title = Ids.text
                  if Ids.tag == 'authorString':
                      child_authorString = Ids.text
                  if Ids.tag == 'journalAbbreviation':
                      child_journalAbbreviation = Ids.text

              citedByObj = {'pmid': child_pmid, 'pubYear': child_pubYear, 'title': child_title, 'authorString': child_authorString, 'journalAbbreviation': child_journalAbbreviation}
              citedBy.append(citedByObj)

      if len(citedBy) != 0:
          new_rec_id = NMODescribing_Overall_CitedBy.insert_one({
                               'pmid': pmid,
                               'doi' : doi,
                               'citedBy': citedBy,
                               'title': None,
                               'publishedDate': None,
                               'source': 'europepmc'
                               })

          #update metadata: doi, title, publishedDate
          for doc in describing.find():
              if doc["pmid"] == pmid:
                  NMODescribing_Overall_CitedBy.update_one({"_id": new_rec_id.inserted_id}, { "$set": {
                                                                 "doi": doc["doi"] ,
                                                                 "title": doc["title"],
                                                                 "publishedDate": doc["publishedDate"] } })
      var_url.close()


@app.route('/users/NMOCiting_References', methods=["POST"])
def NMOCiting_References():

    source = request.values.get('source')
    citing = db.NMO_Citing

    res = []
    NMOCiting_References = db.NMOCiting_References
    NMOCiting_References.drop()
    NMO_Type = "NMO_Citing"
    if source == "crossref":
        crossref_Url = 'http://api.crossref.org/works/'
        urls = []
        for doc in citing.find():
            if doc["doi"] != None:
                urls.append(crossref_Url + str(doc["doi"].strip("\u00a0")))

        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        tasks = [fetchReferences_crossref(i, NMO_Type) for i in urls]
        loop.run_until_complete(asyncio.gather(*tasks))
        loop.close()

    elif source == "europepmc":
        europepmcUrl = 'https://www.ebi.ac.uk/europepmc/webservices/rest/MED/'
        urls = []
        for doc in citing.find():
            if doc["pmid"] != None:
                urls.append(europepmcUrl + str(doc["pmid"]) + '/references')

        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        tasks = [fetchReferences_europepmc(i, NMO_Type) for i in urls]
        loop.run_until_complete(asyncio.gather(*tasks))
        loop.close()

    #Return data as Json
    docs = NMOCiting_References.find()
    for doc in docs:
        doc['_id'] = str(doc['_id'])
        res.append(doc)

    return json.dumps(res)


@app.route('/users/NMOUsing_References', methods=["POST"])
def NMOUsing_References():

    source = request.values.get('source')
    using = db.NMO_Using
    NMOUsing_References = db.NMOUsing_References
    NMOUsing_References.drop()
    NMO_Type = "NMO_Using"

    res = []
    if source == "crossref":
        crossref_Url = 'http://api.crossref.org/works/'
        urls = []
        for doc in using.find():
            if doc["doi"] != None:
                urls.append(crossref_Url + str(doc["doi"].strip("\u00a0")))

        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        tasks = [fetchReferences_crossref(i, NMO_Type) for i in urls]
        loop.run_until_complete(asyncio.gather(*tasks))
        loop.close()

    if source == "europepmc":
        europepmcUrl = 'https://www.ebi.ac.uk/europepmc/webservices/rest/MED/'
        urls = []
        for doc in using.find():
            if doc["pmid"] != None:
                urls.append(europepmcUrl + str(doc["pmid"]) + '/references')

        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        tasks = [fetchReferences_europepmc(i, NMO_Type) for i in urls]
        loop.run_until_complete(asyncio.gather(*tasks))
        loop.close()

    #Return data as Json
    docs = NMOUsing_References.find()
    for doc in docs:
        doc['_id'] = str(doc['_id'])
        res.append(doc)

    return json.dumps(res)


@app.route('/users/NMOCiting_References_Of_Describing', methods=["POST"])
def NMOCiting_References_Of_Describing():

    source = request.values.get('source')

    NMO_Describing = db.NMO_Describing

    res = []
    NMOCiting_References = db.NMOCiting_References
    NMOCiting_References_Of_Describing = db.NMOCiting_References_Of_Describing
    NMOCiting_References_Of_Describing.drop()

    if source == "crossref":
        for doc in NMOCiting_References.find():
            references = doc['references']
            DescribingReferences = []
            for elem in references:
                if elem['doi'] != None:
                    elem_doi = elem['doi']
                    res = NMO_Describing.find_one({'doi': elem_doi})
                    if res != None:
                        Describing_doi = res['doi']
                        if elem_doi == Describing_doi:
                            elem_pmid = res['pmid']
                            elem_title = res['title']
                            elem_publishedDate = res['publishedDate']
                            DescribingReferences.append({'pmid': elem_pmid, 'doi': elem_doi, 'title': elem_title, 'publishedDate': elem_publishedDate})

            if len(DescribingReferences) != 0:
                new_rec_id = NMOCiting_References_Of_Describing.insert_one({
                                                'pmid': doc['pmid'],
                                                'doi' : doc['doi'],
                                                'NMODescribing_References': DescribingReferences,
                                                'title': doc['title'],
                                                'publishedDate': doc['publishedDate'],
                                                'source': doc['source']
                                                 })

    elif source == "europepmc":
        for doc in NMOCiting_References.find():
            references = doc['references']
            DescribingReferences = []
            for elem in references:
                if elem['pmid'] != None:
                    elem_pmid = elem['pmid']
                    res = NMO_Describing.find_one({'pmid': elem_pmid})
                    if res != None:
                        Describing_pmid = res['pmid']
                        if elem_pmid == Describing_pmid:
                            elem_doi = res['doi']
                            elem_title = res['title']
                            elem_publishedDate = res['publishedDate']
                            DescribingReferences.append({'pmid': elem_pmid, 'doi': elem_doi, 'title': elem_title, 'publishedDate': elem_publishedDate})

            if len(DescribingReferences) != 0:
                new_rec_id = NMOCiting_References_Of_Describing.insert_one({
                                                'pmid': doc['pmid'],
                                                'doi' : doc['doi'],
                                                'NMODescribing_References': DescribingReferences,
                                                'title': doc['title'],
                                                'publishedDate': doc['publishedDate'],
                                                'source': doc['source']
                                                 })

    res = []
    docs = NMOCiting_References_Of_Describing.find()
    for doc in docs:
        doc['_id'] = str(doc['_id'])
        res.append(doc)

    return json.dumps(res)


@app.route('/users/NMOUsing_References_Of_Describing', methods=["POST"])
def NMOUsing_References_Of_Describing():

    source = request.values.get('source')

    NMO_Describing = db.NMO_Describing

    res = []
    if source == "crossref":
        NMOUsing_References = db.NMOUsing_References
        NMOUsing_References_Of_Describing = db.NMOUsing_References_Of_Describing

        NMOUsing_References_Of_Describing.drop()

        for doc in NMOUsing_References.find():

            references = doc['references']
            DescribingReferences = []
            for elem in references:
                if elem['doi'] != None:
                    elem_doi = elem['doi']
                    res = NMO_Describing.find_one({'doi': elem_doi})
                    if res != None:
                        Describing_doi = res['doi']
                        if elem_doi == Describing_doi:
                            elem_pmid = res['pmid']
                            elem_title = res['title']
                            elem_publishedDate = res['publishedDate']
                            DescribingReferences.append({'pmid': elem_pmid, 'doi': elem_doi, 'title': elem_title, 'publishedDate': elem_publishedDate})

            if len(DescribingReferences) != 0:
                new_rec_id = NMOUsing_References_Of_Describing.insert_one({
                                                'pmid': doc['pmid'],
                                                'doi' : doc['doi'],
                                                'NMODescribing_References': DescribingReferences,
                                                'title': doc['title'],
                                                'publishedDate': doc['publishedDate'],
                                                'source': doc['source']
                                                 })


    if source == "europepmc":
        NMOUsing_References = db.NMOUsing_References
        NMOUsing_References_Of_Describing = db.NMOUsing_References_Of_Describing

        NMOUsing_References_Of_Describing.drop()

        for doc in NMOUsing_References.find():

            references = doc['references']
            DescribingReferences = []
            for elem in references:
                if elem['pmid'] != None:
                    elem_pmid = elem['pmid']
                    res = NMO_Describing.find_one({'pmid': elem_pmid})
                    if res != None:
                        Describing_pmid = res['pmid']
                        if elem_pmid == Describing_pmid:
                            elem_doi = res['doi']
                            elem_title = res['title']
                            elem_publishedDate = res['publishedDate']
                            DescribingReferences.append({'pmid': elem_pmid, 'doi': elem_doi, 'title': elem_title, 'publishedDate': elem_publishedDate})

            if len(DescribingReferences) != 0:
                new_rec_id = NMOUsing_References_Of_Describing.insert_one({
                                                'pmid': doc['pmid'],
                                                'doi' : doc['doi'],
                                                'NMODescribing_References': DescribingReferences,
                                                'title': doc['title'],
                                                'publishedDate': doc['publishedDate'],
                                                'source': doc['source']
                                                 })

    res = []
    docs = NMOUsing_References_Of_Describing.find()
    for doc in docs:
        doc['_id'] = str(doc['_id'])
        res.append(doc)

    return json.dumps(res)

@app.route('/users/NMODescribing_Overall_CitedBy', methods=["POST"])
def NMODescribing_Overall_CitedBy():

    #source = request.values.get('source')

    NMO_Describing = db.NMO_Describing
    NMODescribing_Overall_CitedBy = db.NMODescribing_Overall_CitedBy
    NMODescribing_Overall_CitedBy.drop()

    #if source == "europepmc":
    europepmcUrl = 'https://www.ebi.ac.uk/europepmc/webservices/rest/MED/'
    urls = []
    for doc in NMO_Describing.find():
        if doc["pmid"] != None:
            urls.append(europepmcUrl + str(doc["pmid"]) + '/citations')

    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    tasks = [fetchCitedBy_europepmc(i) for i in urls]
    loop.run_until_complete(asyncio.gather(*tasks))
    loop.close()

    res = []
    docs = NMODescribing_Overall_CitedBy.find()
    for doc in docs:
        doc['_id'] = str(doc['_id'])
        res.append(doc)

    return json.dumps(res)


@app.route('/users/NMODescribing_Overall_CitedByNMO', methods=["POST"])
def NMODescribing_Overall_CitedByNMO():

    #source = request.values.get('source')

    NMO_Describing = db.NMO_Describing
    NMO_Citing = db.NMO_Citing
    NMO_Using = db.NMO_Using
    NMODescribing_Overall_CitedBy = db.NMODescribing_Overall_CitedBy
    NMODescribing_Overall_CitedByNMO = db.NMODescribing_Overall_CitedByNMO
    NMODescribing_Overall_CitedByNMO.drop()

    #if source == "europepmc":
    docs = NMODescribing_Overall_CitedBy.find()
    for doc in docs:
        citedBy = []
        for elem in doc['citedBy']:
            try:
                if elem['pmid'] != None:
                    NMO_Type = None
                    pmid_link = None
                    doi_link = None
                    doi = None
                    doc_citing = NMO_Citing.find_one({'pmid': elem['pmid']})
                    doc_using  = NMO_Using.find_one({'pmid': elem['pmid']})
                    if doc_citing != None and doc_using != None:
                        NMO_Type = 'NMOCiting_And_NMOUsing'
                    elif doc_citing != None:
                        NMO_Type = 'NMOCiting'
                    elif doc_using != None:
                        NMO_Type = 'NMOUsing'

                    if NMO_Type != None:
                        citedByObj = {}
                        child_pmid_link = 'https://pubmed.ncbi.nlm.nih.gov/' + elem['pmid']
                        child_doi = get_doi(elem['pmid'])
                        child_pubDate = get_pubDate(elem['pmid'])
                        if child_doi != None:
                            child_doi_link = 'https://doi.org/' + child_doi
                        citedByObj = {'pmid': elem['pmid'], 'doi': child_doi, 'pubDate': child_pubDate, 'pubYear': elem['pubYear'],  'title': elem['title'], 'authorString': elem['authorString'], 
                                      'pmid_link': child_pmid_link, 'doi_link': child_doi_link, 'journalAbbreviation': elem['journalAbbreviation'], 'NMO_Type': NMO_Type}
                        citedBy.append(citedByObj)

            except KeyError: pass

        if len(citedBy) != 0:
            total_NMO_Citations = len(citedBy)
            new_rec_id = NMODescribing_Overall_CitedByNMO.insert_one({
                                                'pmid': doc['pmid'],
                                                'doi' : doc['doi'],
                                                'citedBy': citedBy,
                                                'title': doc['title'],
                                                'publishedDate': doc['publishedDate'],
                                                'total_NMO_Citations': total_NMO_Citations
                                                 })

    res = []
    docs = NMODescribing_Overall_CitedByNMO.find()
    for doc in docs:
        doc['_id'] = str(doc['_id'])
        res.append(doc)

    return json.dumps(res)

@app.route('/users/NMODescribing_CitedBy_SinceUpload', methods=["POST"])
def NMODescribing_CitedBy_SinceUpload():

    source = request.values.get('source')

    NMODescribing_CitedBy_SinceUpload = db.NMODescribing_CitedBy_SinceUpload
    NMODescribing_CitedBy_SinceUpload.drop()
    NMODescribing_Overall_CitedBy = db.NMODescribing_Overall_CitedBy
    NMODescribing_UploadDate = db.NMODescribing_UploadDate

    if source == "europepmc":
        fetch_upload_date()
        docs = NMODescribing_Overall_CitedBy.find(no_cursor_timeout=True).batch_size(30)

        recCount = 0
        for doc in docs:
            recCount += 1
            print (recCount)
            #Get Overall CitedBy for each ID
            overallCitedBy = None
            upload_date = None
            bfound = False
            if doc['pmid'] != None:
                #overallCitedBy = getDescribingOverallCitedBy(doc['doi'])
                UploadDate_docs = NMODescribing_UploadDate.find(no_cursor_timeout=True).batch_size(50)
                for elem in UploadDate_docs:
                    if bfound == False:
                        try:
                            for i in range (0, len(elem['reference_doi'])):
                                if doc['doi'] == elem['reference_doi'][i]:
                                    bfound = True
                                    upload_date = elem['upload_date']
                                    print ('subscribing doi:', doc['doi'])
                                    print ('reference_doi:', elem['reference_doi'][i])
                                    print ('date_of_upload:', elem['upload_date'])
                                    break
                        except: pass
                    elif bfound == True:
                        UploadDate_docs.close()
                        break

                UploadDate_docs.close()

                if upload_date == None:  #if upload_date is null with doi search upload_date using pmid
                    UploadDate_docs = NMODescribing_UploadDate.find(no_cursor_timeout=True).batch_size(50)
                    for elem in UploadDate_docs:
                        if bfound == False:
                            try:
                                for i in range (0, len(elem['reference_pmid'])):
                                    if doc['pmid'] == elem['reference_pmid'][i]:
                                        bfound = True
                                        upload_date = elem['upload_date']
                                        break
                            except: pass
                        elif bfound == True:
                            UploadDate_docs.close()
                            break
                    UploadDate_docs.close()

            pubYears = []
            citedBy = []
            for elem in doc['citedBy']:
                try:
                    if upload_date != None:
                        elem_pubYear = elem['pubYear']
                        print ('pubYear:', elem_pubYear)
                        print ('upload_date:', upload_date[:4])
                        ipubYear = int(elem_pubYear)
                        iuploadYear = int(upload_date[:4])

                        print ('pubYear:', ipubYear)
                        print ('uploadYear:', iuploadYear)
                        if  ipubYear >= iuploadYear:
                            print('pubYear >= upload_date')
                            doi = get_doi(elem['pmid'])
                            citedBy.append({'pmid': elem['pmid'], 'doi': doi, 'pubYear': elem['pubYear'], 'title': elem['title'],'authorString': elem['authorString'], 
                                            'journalAbbreviation': elem['journalAbbreviation'], })
                            pubYears.append(elem_pubYear)

                except KeyError: pass


            if len(pubYears) != 0:
                d = defaultdict(int) # values default to 0
                for year in pubYears:
                    d[year] += 1

            if len(citedBy) != 0:
                new_rec_id = NMODescribing_CitedBy_SinceUpload.insert_one({
                                                'pmid': doc['pmid'],
                                                'doi' : doc['doi'],
                                                'citedBy': citedBy,
                                                'title': doc['title'],
                                                'pubYear': doc['publishedDate'][:4],
                                                'upload_date': upload_date,
                                                'NumCitationsByYear_since_upload': dict(d),
                                                'Citations_since_upload': len(citedBy),
                                                'Overall_Citations': len(doc['citedBy'])
                                                 })


    if source == "crossref":
        #fetch_upload_date()
        docs = NMODescribing_ReferencedBy.find(no_cursor_timeout=True).batch_size(30)

        recCount = 0
        for doc in docs:
            recCount += 1
            print ('recCount:', recCount)
            #Get Overall CitedBy for each ID
            overallCitedBy = None
            upload_date = None
            bfound = False
            if doc['doi'] != None:
                #overallCitedBy = getDescribingOverallCitedBy(doc['doi'])
                UploadDate_docs = NMODescribing_UploadDate.find(no_cursor_timeout=True).batch_size(50)
                for elem in UploadDate_docs:
                    if bfound == False:
                        try:
                            for i in range (0, len(elem['reference_doi'])):
                                if doc['doi'] == elem['reference_doi'][i]:
                                    bfound = True
                                    upload_date = elem['upload_date']
                                    print ('describing doi:', doc['doi'])
                                    print ('reference_doi:', elem['reference_doi'][i])
                                    print ('date_of_upload:', elem['upload_date'])
                                    break
                        except: pass
                    elif bfound == True:
                        UploadDate_docs.close()
                        break

                UploadDate_docs.close()

                if upload_date == None:  #if upload_date is null with doi search upload_date using pmid
                    UploadDate_docs = NMODescribing_UploadDate.find(no_cursor_timeout=True).batch_size(50)
                    for elem in UploadDate_docs:
                        if bfound == False:
                            try:
                                for i in range (0, len(elem['reference_pmid'])):
                                    if doc['pmid'] == elem['reference_pmid'][i]:
                                        bfound = True
                                        upload_date = elem['upload_date']
                                        break
                            except: pass
                        elif bfound == True:
                            UploadDate_docs.close()
                            break
                    UploadDate_docs.close()

            pubYears = []
            citedBy = []
            for elem in doc['citedBy']:
                try:
                    if upload_date != None:
                        elem_pubYear = elem['pubYear']
                        print ('pubYear:', elem_pubYear)
                        print ('upload_date:', upload_date[:4])
                        ipubYear = int(elem_pubYear)
                        iuploadYear = 2006 #int(upload_date[:4])

                        print ('pubYear:', ipubYear)
                        print ('uploadYear:', iuploadYear)
                        if  ipubYear >= iuploadYear and elem['pmid'] != None:
                            print('pubYear >= upload_date')
                            citedBy.append({'pmid': elem['pmid'], 'doi': elem['doi'], 'pubYear': elem['pubYear'], 'title': elem['title'], 'type': elem['type']})
                            pubYears.append(elem_pubYear)

                except KeyError: pass


            if len(pubYears) != 0:
                d = defaultdict(int) # values default to 0
                for year in pubYears:
                    d[year] += 1

            if len(citedBy) != 0:
                doc  = NMODescribing_CitedBy_SinceAbsoluteUpload.find_one({'doi': doc['doi']})
                if doc != None:
                    Overall_Citations = doc['citedBy']
                    new_rec_id = NMODescribing_CitedBy_SinceUpload.insert_one({
                                                'pmid': doc['pmid'],
                                                'doi' : doc['doi'],
                                                'citedBy': citedBy,
                                                'title': doc['title'],
                                                'publishedDate': doc['publishedDate'],
                                                'upload_date': upload_date,
                                                'NumCitationsByYear_since_upload': dict(d),
                                                'NMO_Citations_since_upload': len(citedBy),
                                                'Overall_Citations': Overall_Citations
                                                 })


    res = []
    docs = NMODescribing_CitedBy_SinceUpload.find()
    for doc in docs:
        doc['_id'] = str(doc['_id'])
        res.append(doc)

    return json.dumps(res)


@app.route('/users/NMODescribing_CitedBy_NMOs_SinceRelativeUpload', methods=["POST"])
def NMODescribing_CitedBy_NMOs_SinceRelativeUpload():

    source = request.values.get('source')

    NMODescribing_CitedBy_NMOs_SinceRelativeUpload = db.NMODescribing_CitedBy_NMOs_SinceRelativeUpload
    NMODescribing_CitedBy_NMOs_SinceRelativeUpload.drop()
    NMODescribing_CitedBy_SinceUpload = db.NMODescribing_CitedBy_SinceUpload
    NMO_Citing = db.NMO_Citing
    NMO_Using = db.NMO_Using

    if source == "europepmc":
        docs = NMODescribing_CitedBy_SinceUpload.find()
        for doc in docs:
            pubYears = []
            citedBy = []
            for elem in doc['citedBy']:
                try:
                    if elem['pmid'] != None:
                        NMO_Type= None
                        doi = None
                        doi_link = None
                        doc_citing = NMO_Citing.find_one({'pmid': elem['pmid']})
                        doc_using  = NMO_Using.find_one({'pmid': elem['pmid']})
                        if doc_citing != None and doc_using != None:
                            NMO_Type = 'NMOCiting_And_NMOUsing'
                        elif doc_citing != None:
                            NMO_Type = 'NMOCiting'
                        elif doc_using != None:
                            NMO_Type = 'NMOUsing'

                        if NMO_Type != None:
                            pubYears.append(elem['pubYear'])
                            doi = get_doi(elem['pmid'])
                            pmid_link = 'https://pubmed.ncbi.nlm.nih.gov/' + elem['pmid']
                            if doi != None:
                                doi_link = 'https://doi.org/' + elem['doi']
                            citedBy.append({'pmid': elem['pmid'], 'doi': doi, 'pubYear': elem['pubYear'],  'title': elem['title'], 'authorString': elem['authorString'], 
                                            'pmid_link': pmid_link, 'doi_link': doi_link, 'journalAbbreviation': elem['journalAbbreviation'], 'NMO_Type': NMO_Type})

                except KeyError: pass

            if len(pubYears) != 0:
                d = defaultdict(int) # values default to 0
                for year in pubYears:
                    d[year] += 1

            NumCitationdByYear = doc['NumCitationsByYear_since_upload']
            if len(citedBy) != 0:
                NMO_Citations_since_upload = len(citedBy)
                Citations_since_upload = doc['Citations_since_upload']
                Citation_proporsion = round((NMO_Citations_since_upload / Citations_since_upload) * 100, 2)
                Citation_increase = None
                if Citations_since_upload != NMO_Citations_since_upload:
                            Citation_increase = round( NMO_Citations_since_upload / (Citations_since_upload - NMO_Citations_since_upload) * 100, 2)
                upload_year = int(doc['upload_date'][:4])
                upload_month = int(doc['upload_date'][5:7])
                upload_day = int(doc['upload_date'][8:10])

                if upload_month > 6 and upload_day >= 1:
                    upload_year += 1

                Years_since_upload = []
                for elem in citedBy:
                    Years_val = int(elem['pubYear']) - upload_year
                    if Years_val > 0:
                        Years_since_upload.append('Year' + str(Years_val))
                    else:
                        Years_since_upload.append('Year' + str(0))

                if len(Years_since_upload) != 0:
                    d_yr = defaultdict(int)
                    for num_yr in Years_since_upload:
                        d_yr[num_yr] += 1

                relative_percent_yrs_since_upload = []
                for key, val in dict(d_yr).items():
                    percentVal = float(int(val) / int(NMO_Citations_since_upload) * 100)
                    relative_percent_yrs_since_upload.append({'yearNum': key, 'relativePercent': percentVal})

                new_rec_id = NMODescribing_CitedBy_NMOs_SinceRelativeUpload.insert_one({
                                                'pmid': doc['pmid'],
                                                'doi' : doc['doi'],
                                                'citedBy': citedBy,
                                                'title': doc['title'],
                                                'pubYear': doc['pubYear'],
                                                'upload_date': doc['upload_date'],
                                                'NumCitationsByYear_since_upload': dict(d),
                                                'NMO_Citations_since_upload': NMO_Citations_since_upload,
                                                'Citations_since_upload': Citations_since_upload,
                                                'Citation_proporsion': Citation_proporsion,
                                                'Citation_increase': Citation_increase,
                                                'upload_year': upload_year,
                                                'years_since_upload': dict(d_yr),
                                                'relative_percent_years_since_upload': relative_percent_yrs_since_upload
                                                 }) 

    res = []
    docs = NMODescribing_CitedBy_NMOs_SinceRelativeUpload.find()
    for doc in docs:
        doc['_id'] = str(doc['_id'])
        res.append(doc)

    return json.dumps(res)


def fetch_upload_date():

    NMODescribing_UploadDate = db.NMODescribing_UploadDate
    NMODescribing_UploadDate.drop()
    upload_date = None
    NMONeuronUrl = 'http://neuromorpho.org/api/neuron'
    urlHandlerNeuron = urllib.request.urlopen(NMONeuronUrl)
    if urlHandlerNeuron.getcode()==200:
        data = urlHandlerNeuron.read()
        jsonData = json.loads(data)
        #print ('keys in jsonData:', jsonData.keys())
        pages = jsonData['page']['totalPages']
        content = jsonData['_embedded']['neuronResources']
        for i in range(0, pages):
            pageUrl = 'http://neuromorpho.org/api/neuron?page=' + str(i)
            pageUrlHandler = urllib.request.urlopen(pageUrl)
            pageData = pageUrlHandler.read()
            pagejsonData = json.loads(pageData)
            content = pagejsonData['_embedded']['neuronResources']
            print('Page#:', i)
            for j in range(0, len(content)):
                reference_pmid = content[j]['reference_pmid']
                reference_doi = content[j]['reference_doi']
                if reference_pmid != None or reference_doi != None:
                    upload_date = content[j]['upload_date']
                    new_rec_id = NMODescribing_UploadDate.insert_one({
                                                'reference_pmid' : reference_pmid,
                                                'reference_doi' : reference_doi,
                                                'upload_date': upload_date
                                                 })

@app.route('/users/NMODescribingCited', methods=["POST"])
def NMODescribingCited():

    NMO_Describing = db.NMO_Describing
    NMODescribing_Overall_CitedByNMO = db.NMODescribing_Overall_CitedByNMO
    NMO_DescribingCited = db.NMO_DescribingCited
    NMO_DescribingCited.drop()

    docs = NMO_Describing.find()
    for doc in docs:
        pmid = doc['pmid']
        doi = doc['doi']
        doc_pmid = NMODescribing_Overall_CitedByNMO.find_one({'pmid': pmid})
        doc_doi = NMODescribing_Overall_CitedByNMO.find_one({'doi': doi})
        if doc_pmid != None:
            if doc_pmid['pmid'] != None:
                pmid = doc_pmid['pmid'] + str(' *')

        if doc_doi != None:
            if doc_doi['doi'] != None:
                doi = doc_doi['doi'] + str(' *')

        new_rec_id = NMO_DescribingCited.insert_one({
                                                'doi': doi,
                                                'pmid': pmid,
                                                'title': doc['title'],
                                                'publishedDate': doc['publishedDate']
                                                })
    docs.close()

    res = []
    docs = NMO_DescribingCited.find()
    for doc in docs:
        doc['_id'] = str(doc['_id'])
        res.append(doc)

    return json.dumps(res)





@app.route('/users/NMO_Emails', methods=["POST"])

def NMO_Emails():
    email = str(request.values.get('email'))
    IDList = request.get_json()

    alert_IDList = []
    alert_IDsObj = {}
    for elem in IDList:
        alert_IDsObj = {'id': elem['id'].replace(' *', '')}
        alert_IDList.append(alert_IDsObj)

    NMOSubscriptions = db.NMOSubscriptions
    NMODescribing_Overall_CitedByNMO = db.NMODescribing_Overall_CitedByNMO

    s = URLSafeSerializer("secret_key", salt='unsubscribe')
    token = s.dumps(email, salt='unsubscribe')

    data = 'This Email confirms your subscription to get an alert every time the following articles get cited or used: \n' + json.dumps(alert_IDList) + '\n\n'

    describings = []
    for elem in alert_IDList:
        try:
            citedByObj = {}
            citedBy = []
            docs = NMODescribing_Overall_CitedByNMO.find_one({'pmid': elem['id']})
            if docs == None:
                docs = NMODescribing_Overall_CitedByNMO.find_one({'doi': elem['id']})

            if docs != None:
                data += 'pmid: ' + str(docs['pmid']) + ' / doi: ' + str(docs['doi']) + '\n'
                data += 'title: ' + docs['title'] + '\n'
                data += 'publishedDate: ' + docs['publishedDate'] + '\n\n   '

                print ('doc citedBy:', docs['citedBy'])
                for i in range (0, len(docs['citedBy'])):
                        pmid_ = docs['citedBy'][i]['pmid']
                        doi_ = docs['citedBy'][i]['doi']
                        title_ = docs['citedBy'][i]['title']
                        authors_ = docs['citedBy'][i]['authorString']
                        pmid_link_ = docs['citedBy'][i]['pmid_link']
                        doi_link_ = docs['citedBy'][i]['doi_link']
                        journalReference_ = docs['citedBy'][i]['journalAbbreviation']

                        if pmid_ == None:
                            pmid_ = ''
                        if doi_ == None:
                            doi_ = ''
                        if title_ == None:
                            title_ = ''
                        if authors_ == None:
                            authors_ = ''
                        if pmid_link_ == None:
                            pmid_link_ = ''
                        if doi_link_ == None:
                            doi_link_ = ''
                        if journalReference_ == None:
                            journalReference_ = ''

                        data += 'citedBy: \n'
                        data += '--------\n'
                        data += '> pmid: ' + pmid_ + ' / doi: ' + doi_ + '\n'
                        data += '> title: ' + title_ + '\n'
                        data += '> authors: ' + authors_ + '\n'
                        data += '> pmid link: ' + pmid_link_ + '\n'
                        data += '> doi link: ' + doi_link_ + '\n'
                        data += '> journal reference: ' + journalReference_ + '\n\n'

                data += '=============================================\n\n'

            docs.close()
        except:
            pass

    #url = url_for('confirm_unsubscribe', token=token, _external=True)
    url = 'http://cng-nmo-dev3.orc.gmu.edu:5050/users/confirm_unsubscribe/' + str(token)

    msg = Message("NMO Literature Subscription",
                  sender=("Neuromorpho_Admin", "NMOBiblio@gmail.com"),
                  recipients=[email])

    lines = data.rsplit('\n')

    msg.html = render_template('template.html', data=lines, url=url)
    mail.send(msg)

    doc = NMOSubscriptions.find_one({'Email': email})
    if doc == None:
        new_rec_id = NMOSubscriptions.insert_one({
                                              'Email' : email,
                                              'alert_IDList': alert_IDList
                                            })
    else:
        new_alert_IDList = []
        alert_IDsObj = {}
        for elem in doc['alert_IDList']:
            alert_IDsObj = {'id': elem['id']}
            new_alert_IDList.append(alert_IDsObj)

        #check if Ids already Exist
        for val in alert_IDList:
            found_id = False
            for elem in doc['alert_IDList']:
                if val['id'] == elem['id']:
                    found_id = True
                    break
            if found_id == False:
                new_alert_IDList.append(val)

        NMOSubscriptions.update_one({"_id": doc['_id']}, { "$set": {
                                                                 'alert_IDList': new_alert_IDList
                                                                } })

    res = []
    docs = NMOSubscriptions.find()
    for doc in docs:
        doc['_id'] = str(doc['_id'])
        res.append(doc)

    return json.dumps(res)


@app.route('/users/confirm_unsubscribe/<token>')
def confirm_unsubscribe(token):

    response = Response(status=200)

    s = URLSafeSerializer("secret_key", salt='unsubscribe')
    try:
        email = s.loads(token)
    except BadData:
        # find error
        print ('email error')

    #unsubscribe_url = url_for('unsubscribe', email=email, _external=True)
    unsubscribe_url = 'http://cng-nmo-dev3.orc.gmu.edu:5050/users/unsubscribe/' + str(email)

    return render_template('confirm_unsubscribe.html', url=unsubscribe_url, email=email)


@app.route('/users/unsubscribe/<email>')
def unsubscribe(email):

    print ('Unsubscribing........', email)

    NMOSubscriptions = db.NMOSubscriptions
    doc = NMOSubscriptions.find_one({'Email': email})
    messg1 = ''
    messg2 = ''
    messg3 = ''
    if doc != None:
        print ('doc[id]:', doc['_id'])
        NMOSubscriptions.delete_one({"_id": doc['_id']})
        messg1 = 'You have been unsubscribed from this alert notification.'
        messg2 = 'You will not receive any more messages at the email address: '
        messg3 = ', unless you subscribe again.'
    else:
        messg2 = 'The following email address: '
        messg3 = ' is no longer subscribed to the neuromorpho alert notification'

    return render_template('unsubscribed.html', email=email, messg1=messg1, messg2=messg2, messg3=messg3)


@app.route('/users/PreRelease_NMODescribingCitedBy_NMOs', methods=["POST"])
def PreRelease_NMODescribingCitedBy_NMOs():

    NMODescribing_Overall_CitedByNMO = db.NMODescribing_Overall_CitedByNMO
    PreRelease_NMODescribingCitedBy_NMOs = db.PreRelease_NMODescribingCitedBy_NMOs
    PreRelease_NMODescribingCitedBy_NMOs.drop()

    docs = NMODescribing_Overall_CitedByNMO.find()

    for doc in docs:
        new_rec_id = PreRelease_NMODescribingCitedBy_NMOs.insert_one({
                                                    'pmid': doc['pmid'],
                                                    'doi' : doc['doi'],
                                                    'citedBy': doc['citedBy'],
                                                    'title': doc['title'],
                                                    'publishedDate': doc['publishedDate']
                                                    })

    res = []
    docs = PreRelease_NMODescribingCitedBy_NMOs.find()
    for doc in docs:
        doc['_id'] = str(doc['_id'])
        res.append(doc)

    return json.dumps(res)


@app.route('/users/PostRelease_NMODescribingCitedBy_NMOs', methods=["POST"])
def PostRelease_NMODescribingCitedBy_NMOs():

    NMODescribing_Overall_CitedByNMO = db.NMODescribing_Overall_CitedByNMO
    PostRelease_NMODescribingCitedBy_NMOs = db.PostRelease_NMODescribingCitedBy_NMOs
    PostRelease_NMODescribingCitedBy_NMOs.drop()

    docs = NMODescribing_Overall_CitedByNMO.find()

    for doc in docs:
        new_rec_id = PostRelease_NMODescribingCitedBy_NMOs.insert_one({
                               'pmid': doc['pmid'],
                               'doi' : doc['doi'],
                               'citedBy': doc['citedBy'],
                               'title': doc['title'],
                               'publishedDate': doc['publishedDate']
                               })

    res = []
    docs = PostRelease_NMODescribingCitedBy_NMOs.find()
    for doc in docs:
        doc['_id'] = str(doc['_id'])
        res.append(doc)

    return json.dumps(res)


@app.route('/users/Alert_NMODescribingCitedBy_NMOs', methods=["POST"])
def Alert_NMODescribingCitedBy_NMOs():

    PreRelease_NMODescribingCitedBy_NMOs = db.PreRelease_NMODescribingCitedBy_NMOs
    PostRelease_NMODescribingCitedBy_NMOs = db.PostRelease_NMODescribingCitedBy_NMOs
    Alert_NMODescribingCitedBy_NMOs = db.Alert_NMODescribingCitedBy_NMOs
    Alert_NMODescribingCitedBy_NMOs.drop()

    docs_updated = PostRelease_NMODescribingCitedBy_NMOs.find()
    for doc_updated in docs_updated:
        doc_preRelease = PreRelease_NMODescribingCitedBy_NMOs.find_one({'pmid': doc_updated['pmid']})
        citedByObj = {}
        citedBy = []
        if doc_preRelease != None:
            for elem_updated in doc_updated['citedBy']:
                found_match = False
                for elem_preRelease in doc_preRelease['citedBy']:
                    if elem_updated['pmid'] == elem_preRelease['pmid']:
                        found_match = True
                        break
                if found_match == False != None:
                    pubDate = elem_updated['pubDate']
                    if pubDate == None:
                        pubDate = elem_updated['pubYear']
                    citedByObj = {'pmid': elem_updated['pmid'], 'doi': elem_updated['doi'],'pubDate': pubDate,
                                  'title': elem_updated['title'], 'authorString': elem_updated['authorString'], 'pmid_link': elem_updated['pmid_link'],
                                  'doi_link': elem_updated['doi_link'], 'journalAbbreviation': elem_updated['journalAbbreviation']}
                    citedBy.append(citedByObj)

            if len(citedBy) != 0:
                new_rec_id1 = Alert_NMODescribingCitedBy_NMOs.insert_one({
                                                                    'pmid': doc_updated['pmid'],
                                                                    'doi' : doc_updated['doi'],
                                                                    'citedBy': citedBy,
                                                                    'title': doc_updated['title'],
                                                                    'publishedDate': doc_updated['publishedDate']
                                                                    })
        else:
            for elem_updated in doc_updated['citedBy']:
                pubDate = elem_updated['pubDate']
                if pubDate == None:
                        pubDate = elem_updated['pubYear']

                citedByObj = {'pmid': elem_updated['pmid'], 'doi': elem_updated['doi'],'pubDate': elem_updated['pubDate'],
                              'title': elem_updated['title'], 'authorString': elem_updated['authorString'], 'pmid_link': elem_updated['pmid_link'],
                              'doi_link': elem_updated['doi_link'], 'journalAbbreviation': elem_updated['journalAbbreviation']}
                citedBy.append(citedByObj)

            if len(citedBy) != 0:
                new_rec_id2 = Alert_NMODescribingCitedBy_NMOs.insert_one({
                                                                    'pmid': doc_updated['pmid'],
                                                                    'doi' : doc_updated['doi'],
                                                                    'citedBy': citedBy,
                                                                    'title': doc_updated['title'],
                                                                    'publishedDate': doc_updated['publishedDate']
                                                                    })


    docs_updated.close()

    res = []
    docs = Alert_NMODescribingCitedBy_NMOs.find()
    for doc in docs:
        doc['_id'] = str(doc['_id'])
        res.append(doc)

    return json.dumps(res)


@app.route('/users/Trigger_EmailAlert_NMODescribingCitedBy_NMOs', methods=["POST"])
def Trigger_EmailAlert_NMODescribingCitedBy_NMOs():

    response = Response(status=200)

    NMOSubscriptions = db.NMOSubscriptions
    Alert_NMODescribingCitedBy_NMOs = db.Alert_NMODescribingCitedBy_NMOs

    docs_subscribe = NMOSubscriptions.find()
    for doc_subscribe in docs_subscribe:
        email = doc_subscribe['Email']
        alertObj = {}
        alertIDs = []
        data = ''
        for elem_subscribe in doc_subscribe['alert_IDList']:
            docs_alert = Alert_NMODescribingCitedBy_NMOs.find()
            for doc_alert in docs_alert:
                if elem_subscribe['id'] == doc_alert['pmid'] or elem_subscribe['id'] == doc_alert['doi']:
                    alertObj = {'pmid': doc_alert['pmid'], 'doi': doc_alert['doi'], 'title': doc_alert['title'],
                                'publishedDate': doc_alert['publishedDate'], 'citedBy': doc_alert['citedBy']
                                }
                    alertIDs.append(alertObj)
            docs_alert.close()

        if len(alertIDs) != 0:
            print('Email: ', email, 'len: ', len(alertIDs))

            s = URLSafeSerializer("secret_key", salt='unsubscribe')
            token = s.dumps(email, salt='unsubscribe')
            url = url_for('confirm_unsubscribe', token=token, _external=True)

            data = 'This is an alert for the following article(s) getting cited or used: \n'
            for i in range (0, len(alertIDs)):
                pmid = alertIDs[i]['pmid']
                if pmid == None:
                    pmid = ''
                doi = alertIDs[i]['doi']
                if doi == None:
                    doi = ''
                title = alertIDs[i]['title']
                if title == None:
                    title = ''
                publishedDate = alertIDs[i]['publishedDate']
                if publishedDate == None:
                    publishedDate = ''
                citedBy = alertIDs[i]['citedBy']

                data += '\npmid: ' + pmid + '/ doi: ' + doi + '\n'
                data += 'title: ' + title + '\n'
                data += 'publishedDate: ' + publishedDate + '\n\n'

                for j in range (0, len(alertIDs[i]['citedBy'])):
                    pmid_ = alertIDs[i]['citedBy'][j]['pmid']
                    if pmid_ == None:
                        pmid_ = ''
                    doi_ = alertIDs[i]['citedBy'][j]['doi']
                    if doi_ == None:
                        doi_ = ''
                    title_ = alertIDs[i]['citedBy'][j]['title']
                    if title_ == None:
                        title_ = ''
                    pubDate_ = alertIDs[i]['citedBy'][j]['pubDate']
                    if pubDate_ == None:
                        pubDate_ = ''
                    authors_ = alertIDs[i]['citedBy'][j]['authorString']
                    if authors_ == None:
                        authors_ = ''
                    pmid_link_ = alertIDs[i]['citedBy'][j]['pmid_link']
                    if pmid_link_ == None:
                        pmid_link_ = ''
                    doi_link_ = alertIDs[i]['citedBy'][j]['doi_link']
                    if doi_link_ == None:
                        doi_link_ = ''
                    journalReference_ = alertIDs[i]['citedBy'][j]['journalAbbreviation']
                    if journalReference_ == None:
                        journalReference_ = ''
                    data += 'citedBy: \n'
                    data += '--------\n'
                    data += '> pmid: ' + pmid_ + ' doi: ' + doi_ + '\n'
                    data += '> title: ' + title_ + '\n'
                    data += '> pub date: ' + pubDate_ + '\n'
                    data += '> authors: ' + authors_ + '\n'
                    data += '> pmid link: ' + pmid_link_ + '\n'
                    data += '> doi link: ' + doi_link_ + '\n'
                    data += '> journal reference: ' + journalReference_ + '\n\n'

                data += '=============================================\n\n'

            print (data)
            #data = 'This is an alert of the following article(s) getting cited or used: \n' + json.dumps(alertIDs)
            msg = Message("NMO Literature Subscription",
            sender=("Neuromorpho_Admin", "NMOBiblio@gmail.com"),
            recipients=[email])
            #msg.body = data

            lines = data.rsplit('\n')


            print ('lines: ', lines)
            msg.html = render_template('template.html', data=lines, url=url)

            mail.send(msg)

    docs_subscribe = NMOSubscriptions.find()

    return response

@app.route('/users/webhook', methods=["GET"])
def webhook():  #Webhook service for monthly update of lit release and neuron downloads

    response = Response(status=200)

    webhook_log('webhook initiated')

    #Pre-Release of NMO Literature Describing Publications
    PreRelease_NMODescribingCitedBy_NMOs()
    webhook_log('PreRelease_NMODescribingCitedBy_NMOs() completed')

    neuromorphodata()
    webhook_log('neuromorphodata() completed')

    NMODescribing_Overall_CitedBy()
    webhook_log('NMODescribing_Overall_CitedBy() completed')

    NMODescribing_Overall_CitedByNMO()
    webhook_log('NMODescribing_Overall_CitedByNMO() completed')

    NMODescribingCited()
    webhook_log('NMODescribingCited() completed')

    PostRelease_NMODescribingCitedBy_NMOs()
    webhook_log('PostRelease_NMODescribingCitedBy_NMOs() completed')

    Alert_NMODescribingCitedBy_NMOs()
    webhook_log('Alert_NMODescribingCitedBy_NMOs() completed')

    #Trigger Alert for new citations
    Trigger_EmailAlert_NMODescribingCitedBy_NMOs()
    webhook_log('Trigger_EmailAlert_NMODescribingCitedBy_NMOs() completed')

    #Import table dump of neuron_article & logdownload    
    res1 = import_neuron_article_dump()
    if res1.status_code == 200:
        #webhook_log('import_neuron_article_dump completed')
        res2 = import_logdownload_dump()
        if res2.status_code == 200:
            #webhook_log('import logdownload_dump completed')
            res3 = article_neuron_dowload()
            if res3.status_code == 200:
                #webhook_log('article_neuron_dowload completed')
                res4 = number_of_downloads_per_article()
                if res4.status_code == 200:
                    #webhook_log('number_of_downloads_per_article completed')
                    res5 = NMODescribing_All_CitedBy_NMO_With_Downloads()
                    if res5.status_code == 200:
                        #webhook_log('NMODescribing_All_CitedBy_NMO_With_Downloads completed')
                        res6 = NMO_DescribingCited_With_Downloads()
                        if res6.status_code == 200:
                            #webhook_log('NMO_DescribingCited_With_Downloads completed')
                            webhook_log('Import of neuron_article & logdownload to MongoDB completed')


    return response


def webhook_log(msg_body):

    email1 = 'hemissah@masonlive.gmu.edu'
    msg = Message("NMO webhook log",
                  sender=("Neuromorpho_Admin", "NMOBiblio@gmail.com"),
                  recipients=[email1])

    msg.body = msg_body
    mail.send(msg)


    email2 = 'ascoli@gmu.edu'
    msg = Message("NMO webhook log",
                  sender=("Neuromorpho_Admin", "NMOBiblio@gmail.com"),
                  recipients=[email2])

    msg.body = msg_body
    mail.send(msg)



@app.route('/users/DBCollectionsBackup', methods=["POST"])
def DBCollectionsBackup():

    response = Response(status=200)

    NMODescribing_Overall_CitedByNMO = db.NMODescribing_Overall_CitedByNMO
    NMO_DescribingCited = db.NMO_DescribingCited
    NMO_Describing = db.NMO_Describing
    NMO_Citing = db.NMO_Citing
    NMO_Using = db.NMO_Using

    Backup_NMODescribing_Overall_CitedByNMO = db.Backup_NMODescribing_Overall_CitedByNMO
    Backup_NMO_DescribingCited = db.Backup_NMO_DescribingCited
    Backup_NMO_Describing = db.Backup_NMO_Describing
    Backup_NMO_Citing = db.Backup_NMO_Citing
    Backup_NMO_Using = db.Backup_NMO_Using

    Backup_NMODescribing_Overall_CitedByNMO.drop()
    Backup_NMO_DescribingCited.drop()
    Backup_NMO_Describing.drop()
    Backup_NMO_Citing.drop()
    Backup_NMO_Using.drop()

    docs1 = NMODescribing_Overall_CitedByNMO.find()
    for doc1 in docs1:
        new_rec_id1 = Backup_NMODescribing_Overall_CitedByNMO.insert_one({
                               'pmid': doc1['pmid'],
                               'doi' : doc1['doi'],
                               'citedBy': doc1['citedBy'],
                               'title': doc1['title'],
                               'publishedDate': doc1['publishedDate'],
                               'total_NMO_Citations': doc1['total_NMO_Citations']
                               })

    docs2 = NMO_DescribingCited.find()
    for doc2 in docs2:
        new_rec_id2 = Backup_NMO_DescribingCited.insert_one({
                               'pmid': doc2['pmid'],
                               'doi' : doc2['doi'],
                               'title': doc2['title'],
                               'publishedDate': doc2['publishedDate']
                               })

    docs3 = NMO_Describing.find()
    for doc3 in docs3:
        new_rec_id3 = Backup_NMO_Describing.insert_one({
                               'doi' : doc3['doi'],
                               'pmid': doc3['pmid'],
                               'title': doc3['title'],
                               'publishedDate': doc3['publishedDate']
                               })

    docs4 = NMO_Citing.find()
    for doc4 in docs4:
        new_rec_id4 = Backup_NMO_Citing.insert_one({
                               'doi' : doc4['doi'],
                               'pmid': doc4['pmid'],
                               'title': doc4['title'],
                               'publishedDate': doc4['publishedDate']
                               })

    docs5 = NMO_Using.find()
    for doc5 in docs5:
        new_rec_id5 = Backup_NMO_Using.insert_one({
                               'doi' : doc5['doi'],
                               'pmid': doc5['pmid'],
                               'title': doc5['title'],
                               'publishedDate': doc5['publishedDate']
                               })

    docs1.close()
    docs2.close()
    docs3.close()
    docs4.close()
    docs5.close()

    return response


@app.route('/users/DBCollectionsRestore', methods=["POST"])
def DBCollectionsRestore():

    response = Response(status=200)

    Backup_NMODescribing_Overall_CitedByNMO = db.Backup_NMODescribing_Overall_CitedByNMO
    Backup_NMO_DescribingCited = db.Backup_NMO_DescribingCited
    Backup_NMO_Describing = db.Backup_NMO_Describing
    Backup_NMO_Citing = db.Backup_NMO_Citing
    Backup_NMO_Using = db.Backup_NMO_Using

    NMODescribing_Overall_CitedByNMO = db.NMODescribing_Overall_CitedByNMO
    NMO_DescribingCited = db.NMO_DescribingCited
    NMO_Describing = db.NMO_Describing
    NMO_Citing = db.NMO_Citing
    NMO_Using = db.NMO_Using

    NMODescribing_Overall_CitedByNMO.drop()
    NMO_DescribingCited.drop()
    NMO_Describing.drop()
    NMO_Citing.drop()
    NMO_Using.drop()

    docs1 = Backup_NMODescribing_Overall_CitedByNMO.find()
    for doc1 in docs1:
        new_rec_id1 = NMODescribing_Overall_CitedByNMO.insert_one({
                               'pmid': doc1['pmid'],
                               'doi' : doc1['doi'],
                               'citedBy': doc1['citedBy'],
                               'title': doc1['title'],
                               'publishedDate': doc1['publishedDate'],
                               'total_NMO_Citations': doc1['total_NMO_Citations']
                               })

    docs2 = Backup_NMO_DescribingCited.find()
    for doc2 in docs2:
        new_rec_id2 = NMO_DescribingCited.insert_one({
                               'pmid': doc2['pmid'],
                               'doi' : doc2['doi'],
                               'title': doc2['title'],
                               'publishedDate': doc2['publishedDate']
                               })

    docs3 = Backup_NMO_Describing.find()
    for doc3 in docs3:
        new_rec_id3 = NMO_Describing.insert_one({
                               'doi' : doc3['doi'],
                               'pmid': doc3['pmid'],
                               'title': doc3['title'],
                               'publishedDate': doc3['publishedDate']
                               })

    docs4 = Backup_NMO_Citing.find()
    for doc4 in docs4:
        new_rec_id4 = NMO_Citing.insert_one({
                               'doi' : doc4['doi'],
                               'pmid': doc4['pmid'],
                               'title': doc4['title'],
                               'publishedDate': doc4['publishedDate']
                               })

    docs5 = Backup_NMO_Using.find()
    for doc5 in docs5:
        new_rec_id5 = NMO_Using.insert_one({
                               'doi' : doc5['doi'],
                               'pmid': doc5['pmid'],
                               'title': doc5['title'],
                               'publishedDate': doc5['publishedDate']
                               })

    docs1.close()
    docs2.close()
    docs3.close()
    docs4.close()
    docs5.close()

    return response



@app.route('/users/NMODescribing_Bibliometric', methods=["GET"])
def NMODescribing_Bibliometric():
    source = request.values.get('source')

    book = Workbook()

    if source == "crossref":
        NMODescribing_Bibliometric = db.NMODescribing_Bibliometric

        sheet = book.active

        sheet['A1'] = 'pmid'
        sheet['B1'] = 'doi'
        sheet['C1'] = 'pubYear'
        sheet['D1'] = 'upload_date'
        sheet['E1'] = 'Overall_Citations'
        sheet['F1'] = 'TotalNMO_Citations'
        sheet['G1'] = 'Citation_Increase'
        sheet['H1'] = 'ReferencedBy NMOCiting & Using'
        sheet['I1'] = 'Citations ByYear'

        sheet.column_dimensions['A'].width = 9
        sheet.column_dimensions['B'].width = 35
        sheet.column_dimensions['C'].width = 8
        sheet.column_dimensions['D'].width = 12
        sheet.column_dimensions['E'].width = 16
        sheet.column_dimensions['F'].width = 19
        sheet.column_dimensions['G'].width = 17
        sheet.column_dimensions['H'].width = 31
        sheet.column_dimensions['I'].width = 40

        i=2
        for doc in NMODescribing_Bibliometric.find():
            sheet['A'+str(i)] = str(doc['pmid'])
            sheet['B'+str(i)] = str(doc['doi'])
            sheet['C'+str(i)] = str(doc['pubYear'])
            sheet['D'+str(i)] = str(doc['upload_date'])
            sheet['E'+str(i)] = str(doc['Overall_Citations'])
            sheet['F'+str(i)] = str(doc['TotalNMO_Citations'])
            iTotalNMO_Citations = int(doc['TotalNMO_Citations'])
            iOverall_Citations = int(doc['Overall_Citations'])
            citationIncrease = round((iTotalNMO_Citations / iOverall_Citations) * 100, 1)
            sheet['G'+str(i)] = str(citationIncrease) + str('%')
            sheet['I'+str(i)] = str(doc['NumCitationsByYear'])

            citedBy_ids = doc['citedBy']
            for elem in citedBy_ids:
                sheet['H'+str(i)] = elem['doi']
                i += 1

            i += 1

    elif source == "europepmc":
        NMODescribing_CitedBy_NMOs_SinceRelativeUpload = db.NMODescribing_CitedBy_NMOs_SinceRelativeUpload

        create_stat_sheet(book)
        create_byPerct_sheet(book)
        create_byYear_sheet(book)

    filename = "NeuromorphoBibliometric.xlsx"
    book.save(filename)

    #Evaluate cell values
    #excel = ExcelCompiler(filename=filename)
    #for z in [chr(i) for i in range(ord('J'),ord('Y'))]:
    #    print (z+'1:', excel.evaluate('stat!' + z + '1'))


    res = []
    docs = NMODescribing_CitedBy_NMOs_SinceRelativeUpload.find()
    for doc in docs:
        doc['_id'] = str(doc['_id'])
        res.append(doc)
    docs.close()

    return json.dumps(res)


def create_stat_sheet(book):

    sheet = book.active
    NMODescribing_CitedBy_NMOs_SinceRelativeUpload = db.NMODescribing_CitedBy_NMOs_SinceRelativeUpload
    sheet.title = "stat"
    sheet.freeze_panes = "A5"
    sheet.print_title_rows = '1:4' #freeze row 1 - 4

    sheet['I1'] = 'N'
    sheet['I2'] = 'stdev'
    sheet['I3'] = 'mean'

    sheet['A4'] = 'pmid'
    sheet['B4'] = 'doi'
    sheet['C4'] = 'pubYear'
    sheet['D4'] = 'upload_date'
    sheet['E4'] = 'citations_since_upload'
    sheet['F4'] = 'NMO_citations_since_upload'
    sheet['G4'] = 'citation_proportion'
    sheet['H4'] = '% Citation_increase'
    sheet['I4'] = 'upload_year'

    sheet['J4'] = 0
    sheet['K4'] = 1
    sheet['L4'] = 2
    sheet['M4'] = 3
    sheet['N4'] = 4
    sheet['O4'] = 5
    sheet['P4'] = 6
    sheet['Q4'] = 7
    sheet['R4'] = 8
    sheet['S4'] = 9
    sheet['T4'] = 10
    sheet['U4'] = 11
    sheet['V4'] = 12
    sheet['W4'] = 13
    sheet['X4'] = 14

    sheet.column_dimensions['A'].width = 9
    sheet.column_dimensions['B'].width = 35
    sheet.column_dimensions['C'].width = 8
    sheet.column_dimensions['D'].width = 12
    sheet.column_dimensions['E'].width = 21
    sheet.column_dimensions['F'].width = 27
    sheet.column_dimensions['G'].width = 18
    sheet.column_dimensions['H'].width = 18
    sheet.column_dimensions['I'].width = 12

    sheet.column_dimensions['J'].width = 6
    sheet.column_dimensions['K'].width = 6
    sheet.column_dimensions['L'].width = 6
    sheet.column_dimensions['M'].width = 6
    sheet.column_dimensions['N'].width = 6
    sheet.column_dimensions['O'].width = 6
    sheet.column_dimensions['P'].width = 6
    sheet.column_dimensions['Q'].width = 6
    sheet.column_dimensions['R'].width = 6
    sheet.column_dimensions['S'].width = 6
    sheet.column_dimensions['T'].width = 6
    sheet.column_dimensions['U'].width = 6
    sheet.column_dimensions['V'].width = 6
    sheet.column_dimensions['W'].width = 6
    sheet.column_dimensions['X'].width = 6

    i=5
    docs = NMODescribing_CitedBy_NMOs_SinceRelativeUpload.find()
    for doc in docs:
        sheet['A'+str(i)] = int(doc['pmid'])
        sheet['B'+str(i)] = doc['doi']
        sheet['C'+str(i)] = int(doc['pubYear'])
        sheet['D'+str(i)] = doc['upload_date']
        sheet['E'+str(i)] = int(doc['Citations_since_upload'])
        sheet['F'+str(i)] = int(doc['NMO_Citations_since_upload'])
        sheet['G'+str(i)] = float(doc['Citation_proporsion']) * 0.01
        sheet['H'+str(i)] = float(doc['Citation_increase']) * 0.01
        sheet['I'+str(i)] = int(doc['upload_year'])

        for elem in doc['relative_percent_years_since_upload']:
            if elem['yearNum'] == 'Year0':
                sheet['J'+str(i)] = elem['relativePercent']
            elif elem['yearNum'] == 'Year1':
                sheet['K'+str(i)] = elem['relativePercent']
            elif elem['yearNum'] == 'Year2':
                sheet['L'+str(i)] = elem['relativePercent']
            elif elem['yearNum'] == 'Year3':
                sheet['M'+str(i)] = elem['relativePercent']
            elif elem['yearNum'] == 'Year4':
                sheet['N'+str(i)] = elem['relativePercent']
            elif elem['yearNum'] == 'Year5':
                sheet['O'+str(i)] = elem['relativePercent']
            elif elem['yearNum'] == 'Year6':
                sheet['P'+str(i)] = elem['relativePercent']
            elif elem['yearNum'] == 'Year7':
                sheet['Q'+str(i)] = elem['relativePercent']
            elif elem['yearNum'] == 'Year8':
                sheet['R'+str(i)] = elem['relativePercent']
            elif elem['yearNum'] == 'Year9':
                sheet['S'+str(i)] = elem['relativePercent']
            elif elem['yearNum'] == 'Year10':
                sheet['T'+str(i)] = elem['relativePercent']
            elif elem['yearNum'] == 'Year11':
                sheet['U'+str(i)] = elem['relativePercent']
            elif elem['yearNum'] == 'Year12':
                sheet['V'+str(i)] = elem['relativePercent']
            elif elem['yearNum'] == 'Year13':
                sheet['W'+str(i)] = elem['relativePercent']
            elif elem['yearNum'] == 'Year14':
                sheet['X'+str(i)] = elem['relativePercent']

            currentYear = 2020
            for j in range (10, 10+(currentYear - int(doc['upload_year']))+1):
                cell = sheet.cell(row=i,column=j)
                if cell.value == None:
                    cell.value = 0
                    cell.number_format = numbers.FORMAT_NUMBER
                elif int(cell.value) == cell.value:
                    cell.number_format = numbers.FORMAT_NUMBER
                else:
                    cell.number_format = numbers.FORMAT_NUMBER_00

        i += 1
    max_col = sheet.max_column
    max_row = sheet.max_row
    #set format of first row column(pmid)as numbers
    for j in range(5,max_row+1):
        cell = sheet.cell(row=j,column=1)
        cell.number_format = numbers.FORMAT_NUMBER

    #set format of third column (pubYear) as numbers
    for j in range(5,max_row+1):
        cell = sheet.cell(row=j,column=3)
        cell.number_format = numbers.FORMAT_NUMBER

    #set format of fifth column (citations_since_upload) as numbers
    for j in range(5,max_row+1):
        cell = sheet.cell(row=j,column=5)
        cell.number_format = numbers.FORMAT_NUMBER

    #set format of sixth column (NMO_Citations_since_upload) as numbers
    for j in range(5,max_row+1):
        cell = sheet.cell(row=j,column=6)
        cell.number_format = numbers.FORMAT_NUMBER

    #set format of seventh column (citation_proportion) as percentage
    for j in range(5,max_row+1):
        cell = sheet.cell(row=j,column=7)
        cell.number_format = numbers.FORMAT_PERCENTAGE_00

    #set format of eigth column (citation_increase) as percentage
    for j in range(5,max_row+1):
        cell = sheet.cell(row=j,column=8)
        cell.number_format = numbers.FORMAT_PERCENTAGE_00

    #set format of ninth column (upload_year) as percentage
    for j in range(5,max_row+1):
        cell = sheet.cell(row=j,column=9)
        cell.number_format = numbers.FORMAT_NUMBER

    #determine columns to skip if all counts are zero or have no values
    skipColumn = []
    for k in range (10,max_col+1):
        values = []
        for j in range(5,max_row+1):
            cell = sheet.cell(row=j, column=k)
            if cell.value != None:
                values.append(float(cell.value))

        if sum(values) == 0.00 or len(values) == 0:
            print ('k:', k)
            if k == 10:
                skipColumn.append('J')
            elif k == 11:
                skipColumn.append('K')
            elif k == 12:
                skipColumn.append('L')
            elif k == 13:
                skipColumn.append('M')
            elif k == 14:
                skipColumn.append('N')
            elif k == 15:
                skipColumn.append('O')
            elif k == 16:
                skipColumn.append('P')
            elif k == 17:
                skipColumn.append('Q')
            elif k == 18:
                skipColumn.append('R')
            elif k == 19:
                skipColumn.append('S')
            elif k == 20:
                skipColumn.append('T')
            elif k == 21:
                skipColumn.append('U')
            elif k == 22:
                skipColumn.append('V')
            elif k == 23:
                skipColumn.append('W')
            elif k == 24:
                skipColumn.append('X')

    #calculate Count of cells with percent values
    for z in [chr(i) for i in range(ord('J'),ord('Y'))]:
        skip = False
        for col in skipColumn:
            if col == z:
                skip = True
        if skip == False:
            sheet[z + '1'] = '= COUNT(' + z + '5:' + z + str(max_row) + ')'
            cell = sheet[z + '1']
            cell.number_format = numbers.FORMAT_NUMBER

    #calculate Standard Deviation values
    for z in [chr(i) for i in range(ord('J'),ord('Y'))]:
        skip = False
        for col in skipColumn:
            if col == z:
                skip = True
        if skip == False:
            sheet[z + '2'] = '= STDEV(' + z + '5:' + z + str(max_row) + ')'
            cell = sheet[z + '2']
            cell.number_format = numbers.FORMAT_NUMBER_00

    #calculate Mean values
    for z in [chr(i) for i in range(ord('J'),ord('Y'))]:
        skip = False
        for col in skipColumn:
            if col == z:
                skip = True
        if skip == False:
            sheet[z + '3'] = '= AVERAGE(' + z + '5:' + z + str(max_row) + ')'
            cell = sheet[z + '3']
            cell.number_format = numbers.FORMAT_NUMBER

    docs.close()

def create_byPerct_sheet(book):

    sheet = book.create_sheet(index = 1 , title = "by-Perct")
    NMODescribing_CitedBy_NMOs_SinceRelativeUpload = db.NMODescribing_CitedBy_NMOs_SinceRelativeUpload
    sheet.freeze_panes = "A2"
    sheet.print_title_rows = '1:1' #freeze first row

    sheet['A1'] = 'pmid'
    sheet['B1'] = 'doi'
    sheet['C1'] = 'pubYear'
    sheet['D1'] = 'upload_date'
    sheet['E1'] = 'citations_since_upload'
    sheet['F1'] = 'NMO_citations_since_upload'
    sheet['G1'] = 'citation_proportion'
    sheet['H1'] = '% Citation_increase'
    sheet['I1'] = 'upload_year'

    sheet['J1'] = 0
    sheet['K1'] = 1
    sheet['L1'] = 2
    sheet['M1'] = 3
    sheet['N1'] = 4
    sheet['O1'] = 5
    sheet['P1'] = 6
    sheet['Q1'] = 7
    sheet['R1'] = 8
    sheet['S1'] = 9
    sheet['T1'] = 10
    sheet['U1'] = 11
    sheet['V1'] = 12
    sheet['W1'] = 13
    sheet['X1'] = 14

    sheet.column_dimensions['A'].width = 9
    sheet.column_dimensions['B'].width = 35
    sheet.column_dimensions['C'].width = 8
    sheet.column_dimensions['D'].width = 12
    sheet.column_dimensions['E'].width = 21
    sheet.column_dimensions['F'].width = 27
    sheet.column_dimensions['G'].width = 18
    sheet.column_dimensions['H'].width = 18
    sheet.column_dimensions['I'].width = 12

    sheet.column_dimensions['J'].width = 6
    sheet.column_dimensions['K'].width = 6
    sheet.column_dimensions['L'].width = 6
    sheet.column_dimensions['M'].width = 6
    sheet.column_dimensions['N'].width = 6
    sheet.column_dimensions['O'].width = 6
    sheet.column_dimensions['P'].width = 6
    sheet.column_dimensions['Q'].width = 6
    sheet.column_dimensions['R'].width = 6
    sheet.column_dimensions['S'].width = 6
    sheet.column_dimensions['T'].width = 6
    sheet.column_dimensions['U'].width = 6
    sheet.column_dimensions['V'].width = 6
    sheet.column_dimensions['W'].width = 6
    sheet.column_dimensions['X'].width = 6

    i=2
    docs = NMODescribing_CitedBy_NMOs_SinceRelativeUpload.find()
    for doc in docs:
        sheet['A'+str(i)] = int(doc['pmid'])
        sheet['B'+str(i)] = doc['doi']
        sheet['C'+str(i)] = int(doc['pubYear'])
        sheet['D'+str(i)] = doc['upload_date']
        sheet['E'+str(i)] = int(doc['Citations_since_upload'])
        sheet['F'+str(i)] = int(doc['NMO_Citations_since_upload'])
        sheet['G'+str(i)] = float(doc['Citation_proporsion']) * 0.01
        sheet['H'+str(i)] = float(doc['Citation_increase']) * 0.01
        sheet['I'+str(i)] = int(doc['upload_year'])

        for elem in doc['relative_percent_years_since_upload']:
            if elem['yearNum'] == 'Year0':
                sheet['J'+str(i)] = elem['relativePercent']
            elif elem['yearNum'] == 'Year1':
                sheet['K'+str(i)] = elem['relativePercent']
            elif elem['yearNum'] == 'Year2':
                sheet['L'+str(i)] = elem['relativePercent']
            elif elem['yearNum'] == 'Year3':
                sheet['M'+str(i)] = elem['relativePercent']
            elif elem['yearNum'] == 'Year4':
                sheet['N'+str(i)] = elem['relativePercent']
            elif elem['yearNum'] == 'Year5':
                sheet['O'+str(i)] = elem['relativePercent']
            elif elem['yearNum'] == 'Year6':
                sheet['P'+str(i)] = elem['relativePercent']
            elif elem['yearNum'] == 'Year7':
                sheet['Q'+str(i)] = elem['relativePercent']
            elif elem['yearNum'] == 'Year8':
                sheet['R'+str(i)] = elem['relativePercent']
            elif elem['yearNum'] == 'Year9':
                sheet['S'+str(i)] = elem['relativePercent']
            elif elem['yearNum'] == 'Year10':
                sheet['T'+str(i)] = elem['relativePercent']
            elif elem['yearNum'] == 'Year11':
                sheet['U'+str(i)] = elem['relativePercent']
            elif elem['yearNum'] == 'Year12':
                sheet['V'+str(i)] = elem['relativePercent']
            elif elem['yearNum'] == 'Year13':
                sheet['W'+str(i)] = elem['relativePercent']
            elif elem['yearNum'] == 'Year14':
                sheet['X'+str(i)] = elem['relativePercent']

            currentYear = 2020
            for j in range (10, 10+(currentYear - int(doc['upload_year']))+1):
                cell = sheet.cell(row=i,column=j)
                if cell.value == None:
                    cell.value = 0
                    cell.number_format = numbers.FORMAT_NUMBER
                elif int(cell.value) == cell.value:
                    cell.number_format = numbers.FORMAT_NUMBER
                else:
                    cell.number_format = numbers.FORMAT_NUMBER_00

        i += 1

    max_col = sheet.max_column
    max_row = sheet.max_row
    #set format of first row column(pmid)as numbers
    for j in range(2,max_row+1):
        cell = sheet.cell(row=j,column=1)
        cell.number_format = numbers.FORMAT_NUMBER

    #set format of third column (pubYear) as numbers
    for j in range(2,max_row+1):
        cell = sheet.cell(row=j,column=3)
        cell.number_format = numbers.FORMAT_NUMBER

    #set format of fifth column (citations_since_upload) as numbers
    for j in range(2,max_row+1):
        cell = sheet.cell(row=j,column=5)
        cell.number_format = numbers.FORMAT_NUMBER

    #set format of sixth column (NMO_Citations_since_upload) as numbers
    for j in range(2,max_row+1):
        cell = sheet.cell(row=j,column=6)
        cell.number_format = numbers.FORMAT_NUMBER

    #set format of seventh column (citation_proportion) as percentage
    for j in range(2,max_row+1):
        cell = sheet.cell(row=j,column=7)
        cell.number_format = numbers.FORMAT_PERCENTAGE_00

    #set format of eigth column (citation_increase) as percentage
    for j in range(2,max_row+1):
        cell = sheet.cell(row=j,column=8)
        cell.number_format = numbers.FORMAT_PERCENTAGE_00

    #set format of ninth column (upload_year) as percentage
    for j in range(2,max_row+1):
        cell = sheet.cell(row=j,column=9)
        cell.number_format = numbers.FORMAT_NUMBER

    docs.close()


def create_byYear_sheet(book):

    sheet = book.create_sheet(index = 1 , title = "by-Year")
    NMODescribing_CitedBy_NMOs_SinceRelativeUpload = db.NMODescribing_CitedBy_NMOs_SinceRelativeUpload
    sheet.freeze_panes = "A2"
    sheet.print_title_rows = '1:1' #freeze first row

    sheet['A1'] = 'pmid'
    sheet['B1'] = 'doi'
    sheet['C1'] = 'pubYear'
    sheet['D1'] = 'upload_date'
    sheet['E1'] = 'citations_since_upload'
    sheet['F1'] = 'NMO_citations_since_upload'
    sheet['G1'] = 'citation_proportion'
    sheet['H1'] = '% Citation_increase'
    sheet['I1'] = 'upload_year'

    sheet['J1'] = 0
    sheet['K1'] = 1
    sheet['L1'] = 2
    sheet['M1'] = 3
    sheet['N1'] = 4
    sheet['O1'] = 5
    sheet['P1'] = 6
    sheet['Q1'] = 7
    sheet['R1'] = 8
    sheet['S1'] = 9
    sheet['T1'] = 10
    sheet['U1'] = 11
    sheet['V1'] = 12
    sheet['W1'] = 13
    sheet['X1'] = 14

    sheet.column_dimensions['A'].width = 9
    sheet.column_dimensions['B'].width = 35
    sheet.column_dimensions['C'].width = 8
    sheet.column_dimensions['D'].width = 12
    sheet.column_dimensions['E'].width = 21
    sheet.column_dimensions['F'].width = 27
    sheet.column_dimensions['G'].width = 18
    sheet.column_dimensions['H'].width = 18
    sheet.column_dimensions['I'].width = 12

    sheet.column_dimensions['J'].width = 6
    sheet.column_dimensions['K'].width = 6
    sheet.column_dimensions['L'].width = 6
    sheet.column_dimensions['M'].width = 6
    sheet.column_dimensions['N'].width = 6
    sheet.column_dimensions['O'].width = 6
    sheet.column_dimensions['P'].width = 6
    sheet.column_dimensions['Q'].width = 6
    sheet.column_dimensions['R'].width = 6
    sheet.column_dimensions['S'].width = 6
    sheet.column_dimensions['T'].width = 6
    sheet.column_dimensions['U'].width = 6
    sheet.column_dimensions['V'].width = 6
    sheet.column_dimensions['W'].width = 6
    sheet.column_dimensions['X'].width = 6

    i=2
    docs = NMODescribing_CitedBy_NMOs_SinceRelativeUpload.find()
    for doc in docs:
        sheet['A'+str(i)] = int(doc['pmid'])
        sheet['B'+str(i)] = doc['doi']
        sheet['C'+str(i)] = int(doc['pubYear'])
        sheet['D'+str(i)] = doc['upload_date']
        sheet['E'+str(i)] = int(doc['Citations_since_upload'])
        sheet['F'+str(i)] = int(doc['NMO_Citations_since_upload'])
        sheet['G'+str(i)] = float(doc['Citation_proporsion']) * 0.01
        sheet['H'+str(i)] = float(doc['Citation_increase']) * 0.01
        sheet['I'+str(i)] = int(doc['upload_year'])

        for key, val in dict(doc['years_since_upload']).items():
            if key == 'Year0':
                sheet['J'+str(i)] = val
            elif key == 'Year1':
                sheet['K'+str(i)] = val
            elif key == 'Year2':
                sheet['L'+str(i)] = val
            elif key == 'Year3':
                sheet['M'+str(i)] = val
            elif key == 'Year4':
                sheet['N'+str(i)] = val
            elif key == 'Year5':
                sheet['O'+str(i)] = val
            elif key == 'Year6':
                sheet['P'+str(i)] = val
            elif key == 'Year7':
                sheet['Q'+str(i)] = val
            elif key == 'Year8':
                sheet['R'+str(i)] = val
            elif key == 'Year9':
                sheet['S'+str(i)] = val
            elif key == 'Year10':
                sheet['T'+str(i)] = val
            elif key == 'Year11':
                sheet['U'+str(i)] = val
            elif key == 'Year12':
                sheet['V'+str(i)] = val
            elif key == 'Year13':
                sheet['W'+str(i)] = val
            elif key == 'Year14':
                sheet['X'+str(i)] = val

            currentYear = 2020
            for j in range (10, 10+(currentYear - int(doc['upload_year']))+1):
                cell = sheet.cell(row=i,column=j)
                if cell.value == None:
                    cell.value = 0
                cell.number_format = numbers.FORMAT_NUMBER

        i += 1

    max_col = sheet.max_column
    max_row = sheet.max_row
    #set format of first row column(pmid)as numbers
    for j in range(2,max_row+1):
        cell = sheet.cell(row=j,column=1)
        cell.number_format = numbers.FORMAT_NUMBER

    #set format of third column (pubYear) as numbers
    for j in range(2,max_row+1):
        cell = sheet.cell(row=j,column=3)
        cell.number_format = numbers.FORMAT_NUMBER

    #set format of fifth column (citations_since_upload) as numbers
    for j in range(2,max_row+1):
        cell = sheet.cell(row=j,column=5)
        cell.number_format = numbers.FORMAT_NUMBER

    #set format of sixth column (NMO_Citations_since_upload) as numbers
    for j in range(2,max_row+1):
        cell = sheet.cell(row=j,column=6)
        cell.number_format = numbers.FORMAT_NUMBER

    #set format of seventh column (citation_proportion) as percentage
    for j in range(2,max_row+1):
        cell = sheet.cell(row=j,column=7)
        cell.number_format = numbers.FORMAT_PERCENTAGE_00

    #set format of eigth column (citation_increase) as percentage
    for j in range(2,max_row+1):
        cell = sheet.cell(row=j,column=8)
        cell.number_format = numbers.FORMAT_PERCENTAGE_00

    #set format of ninth column (upload_year) as percentage
    for j in range(2,max_row+1):
        cell = sheet.cell(row=j,column=9)
        cell.number_format = numbers.FORMAT_NUMBER

    #set format of tenth column thru last column as number
    for j in range (10, max_col+1):
        for k in range(2, max_row+1):
            cell = sheet.cell(row=k,column=j)
            if cell.value != None:
                cell.number_format = numbers.FORMAT_NUMBER

    docs.close()

@app.route('/users/NMODescribing', methods=["GET"])
def NMODescribing():
    source = request.values.get('source')
    res = []
    if source == "europepmc":
        NMO_Describing = db.NMO_Describing
        docs = NMO_Describing.find()
        for doc in docs:
            doc['_id'] = str(doc['_id'])
            res.append(doc)

    return json.dumps(res)


@app.route('/users/NMODescribingCitedList', methods=["GET"])
def NMODescribingCitedList():
    source = request.values.get('source')
    res = []
    if source == "europepmc":
        #NMO_DescribingCited = db.NMO_DescribingCited
        #docs = NMO_DescribingCited.find()
        NMO_DescribingCited_With_Downloads = db.NMO_DescribingCited_With_Downloads
        docs = NMO_DescribingCited_With_Downloads.find()
        for doc in docs:
            doc['_id'] = str(doc['_id'])
            res.append(doc)

    return json.dumps(res)


@app.route('/users/NMODescribing_All_CitedBy', methods=["GET"])
def NMODescribing_All_CitedBy():
    source = request.values.get('source')
    res = []
    if source == "europepmc":
        NMODescribing_Overall_CitedBy = db.NMODescribing_Overall_CitedBy
        docs = NMODescribing_Overall_CitedBy.find()
        for doc in docs:
            doc['_id'] = str(doc['_id'])
            res.append(doc)

    return json.dumps(res)

@app.route('/users/NMODescribing_All_CitedBy_NMOs', methods=["GET"])
def NMODescribing_All_CitedBy_NMOs():
    source = request.values.get('source')
    res = []
    if source == "europepmc":
        NMODescribing_Overall_CitedByNMO = db.NMODescribing_Overall_CitedByNMO
        docs = NMODescribing_Overall_CitedByNMO.find()
        for doc in docs:
            doc['_id'] = str(doc['_id'])
            res.append(doc)

    return json.dumps(res)

@app.route('/users/NMODescribing_All_CitedBy_SinceUpload', methods=["GET"])
def NMODescribing_All_CitedBy_SinceUpload():
    source = request.values.get('source')
    res = []
    if source == "europepmc":
        NMODescribing_CitedBy_SinceUpload = db.NMODescribing_CitedBy_SinceUpload
        docs = NMODescribing_CitedBy_SinceUpload.find()
        for doc in docs:
            doc['_id'] = str(doc['_id'])
            res.append(doc)

    return json.dumps(res)

@app.route('/users/NMODescribing_CitedBy_NMOs_SinceUpload', methods=["GET"])
def NMODescribing_CitedBy_NMOs_SinceUpload():
    source = request.values.get('source')
    res = []
    if source == "europepmc":
        NMODescribing_CitedBy_NMOs_SinceRelativeUpload = db.NMODescribing_CitedBy_NMOs_SinceRelativeUpload
        docs = NMODescribing_CitedBy_NMOs_SinceRelativeUpload.find()
        for doc in docs:
            doc['_id'] = str(doc['_id'])
            res.append(doc)

    return json.dumps(res)

@app.route('/users/NMODescribing_Overall_CitedBy_ByID', methods=["GET"])
def NMODescribing_Overall_CitedBy_ByID():
    source = request.values.get('source')
    pmid = request.values.get('pmid')
    doi = request.values.get('doi')

    res = []
    if source == "europepmc":
        NMODescribing_Overall_CitedBy = db.NMODescribing_Overall_CitedBy
        doc = None
        if pmid != None:
            doc = NMODescribing_Overall_CitedBy.find_one({'pmid': pmid})
        elif doi != None:
            doc = NMODescribing_Overall_CitedBy.find_one({'doi': doi})

        if doc != None:
            res = ({'pmid': doc['pmid'], 'doi': doc['doi'], 'citedBy': doc['citedBy'],
                    'title': doc['title'], 'publishedDate': doc['publishedDate'], 'source': doc['source']})

    return json.dumps(res)


@app.route('/users/NMODescribing_All_CitedBy_SinceUpload_ByID', methods=["GET"])
def NMODescribing_All_CitedBy_SinceUpload_ByID():
    source = request.values.get('source')
    pmid = request.values.get('pmid')
    doi = request.values.get('doi')

    res = []
    if source == "europepmc":
        NMODescribing_CitedBy_SinceUpload = db.NMODescribing_CitedBy_SinceUpload
        doc = None
        if pmid != None:
            doc = NMODescribing_CitedBy_SinceUpload.find_one({'pmid': pmid})
        elif doi != None:
            doc = NMODescribing_CitedBy_SinceUpload.find_one({'doi': doi})

        if doc != None:
            res = ({'pmid': doc['pmid'], 'doi': doc['doi'], 'citedBy': doc['citedBy'], 'title': doc['title'],
                    'pubYear': doc['pubYear'], 'upload_date': doc['upload_date'],
                    'NumCitationsByYear_since_upload': doc['NumCitationsByYear_since_upload'],
                    'Citations_since_upload': doc['Citations_since_upload'], 'Overall_Citations': doc['Overall_Citations']})

    return json.dumps(res)

@app.route('/users/NMODescribing_Overall_CitedByNMO_ByID', methods=["GET"])
def NMODescribing_Overall_CitedByNMO_ByID():
    source = request.values.get('source')
    pmid = request.values.get('pmid')
    doi = request.values.get('doi')

    res = []
    if source == "europepmc":
        #NMODescribing_Overall_CitedByNMO = db.NMODescribing_Overall_CitedByNMO
        NMODescribing_Overall_CitedByNMO_With_Downloads = db.NMODescribing_Overall_CitedByNMO_With_Downloads
        doc = None
        fullurl=None
        if pmid != None:
            doc = NMODescribing_Overall_CitedByNMO_With_Downloads.find_one({'pmid': pmid})
            if doc != None:
                res = ({'pmid': doc['pmid'], 'doi': doc['doi'], 'citedBy': doc['citedBy'], 'title': doc['title'],
                       'publishedDate': doc['publishedDate'], 'total_NMO_Citations': doc['total_NMO_Citations'],
                       'Number_of_Downloads': doc['Number_of_Downloads'], 'Avg_Downloads_Per_Cell': doc['Avg_Downloads_Per_Cell']})
        elif doi != None:
           fullurl = request.url

           doistart = fullurl.find("?doi=")
           doiend = fullurl.find("&source=")

           if doistart < 0:
               doistart = fullurl.find("&doi=")
               doiend = len(fullurl)

           #symbol & is reserved in parameter split - replace & in url with chr(38)
           doi = urllib.parse.unquote(fullurl[doistart+5 : doiend])
           querydoi = doi.replace('&', chr(38))
           filter={'doi': querydoi}

           doc = NMODescribing_Overall_CitedByNMO_With_Downloads.find(filter=filter)
           for elem in doc:
               res = ({'pmid': elem['pmid'], 'doi': elem['doi'], 'citedBy': elem['citedBy'], 'title': elem['title'], 
                       'publishedDate': elem['publishedDate'], 'total_NMO_Citations': elem['total_NMO_Citations'],
                       'Number_of_Downloads': elem['Number_of_Downloads'], 'Avg_Downloads_Per_Cell': elem['Avg_Downloads_Per_Cell']})


    return json.dumps(res)


@app.route('/users/NMODescribing_CitedBy_NMOs_SinceUpload_ByID', methods=["GET"])
def NMODescribing_CitedBy_NMOs_SinceUpload_ByID():
    source = request.values.get('source')
    pmid = request.values.get('pmid')
    doi = request.values.get('doi')

    res = []
    if source == "europepmc":
        NMODescribing_CitedBy_NMOs_SinceRelativeUpload = db.NMODescribing_CitedBy_NMOs_SinceRelativeUpload
        doc = None
        if pmid != None:
            doc = NMODescribing_CitedBy_NMOs_SinceRelativeUpload.find_one({'pmid': pmid})
        elif doi != None:
            doc = NMODescribing_CitedBy_NMOs_SinceRelativeUpload.find_one({'doi': doi})

        if doc != None:
            res = ({'pmid': doc['pmid'], 'doi': doc['doi'], 'citedBy': doc['citedBy'], 'title': doc['title'],
                    'pubYear': doc['pubYear'], 'upload_date': doc['upload_date'],
                    'NumCitationsByYear_since_upload': doc['NumCitationsByYear_since_upload'],
                    'NMO_Citations_since_upload': doc['NMO_Citations_since_upload'],
                    'Citations_since_upload': doc['Citations_since_upload'], 'Citation_proporsion': doc['Citation_proporsion'],
                    'Citation_increase': doc['Citation_increase'], 'upload_year': doc['upload_year'],
                    'years_since_upload': doc['years_since_upload'], 'relative_percent_years_since_upload': doc['relative_percent_years_since_upload']})

    return json.dumps(res)


@app.route('/users/import_neuron_article_dump', methods=['POST'])
def import_neuron_article_dump():

    NMO_Neuron_Article = db.NMO_Neuron_Article
    NMO_Neuron_Article.drop()

    dump_dir = '/nmo/tables_dump'
    if not os.path.exists(dump_dir):
       os.makedirs(dump_dir)

    os.chdir(dump_dir)

    sql_file = 'neuron_article.sql'
    if not os.path.exists(sql_file):
        print(f"File {sql_file} does not exist.")
    else:
        print(f"File {sql_file} exists.")

    try:
        # Read the SQL dump file
        with open(sql_file, 'r') as sql_file:
            queries  = sql_file.read()

        # Process and insert data into MongoDB
        data = []
        insert_pattern = re.compile(r"INSERT INTO `neuron_article` VALUES \((.*?)\);", re.MULTILINE | re.DOTALL)
        insert_matches = re.findall(insert_pattern, queries)

        for insert_match in insert_matches:
            records = insert_match.split("),")
            for record in records:
                values = [int(v) if v.isdigit() else v.strip("'") if v != "NULL" else None for v in record.split(",")]

                # Remove open parenthesis from index_id if it exists
                if str(values[0]).find('(') == 0:
                    values[0] = values[0].replace('(', '')

                datum = {
                    'neuron_id': int(values[0]),
                    'article_id': values[1],
                    'PMID': values[2],
                    'id': values[3]
                }
                data.append(datum)

        result = NMO_Neuron_Article.insert_many(data)
        if result.acknowledged:
            response_data = {'message': 'Data imported successfully'}
            response = make_response(jsonify(response_data))
            response.status_code = 200
            return response
        else:
            response_data = {'error': 'Data import failed'}
            return jsonify(response_data), 500

    except Exception as e:
        print('An error occurred:', str(e))
        return abort(500, description='An error occurred')

@app.route('/users/import_logdownload_dump', methods=['POST'])
def import_logdownload_dump():

    NMO_Log_Download = db.NMO_Log_Download
    NMO_Log_Download.drop()

    dump_dir = '/nmo/tables_dump'
    if not os.path.exists(dump_dir):
       os.makedirs(dump_dir)

    os.chdir(dump_dir)

    sql_file = 'logdownload.sql'
    if not os.path.exists(sql_file):
        print(f"File {sql_file} does not exist.")
    else:
        print(f"File {sql_file} exists.")

    with open(sql_file, 'r', encoding='utf-8') as file:
      try:
        insert_pattern = re.compile(r"INSERT INTO `logdownload` VALUES (.*?);", re.MULTILINE | re.DOTALL)

        batch_size = 1000

        records_to_insert = []

        for line in file:
            insert_matches = re.findall(insert_pattern, line)

            for insert_match in insert_matches:
                records = insert_match.split("),")
                for record in records:
                    values = [v.strip("'") if v != "NULL" else None for v in re.split(r",(?![^(]*\))", record)]

                    if len(values) >= 11:
                        columns_to_keep_as_strings = [1, 2, 3, 4, 10]
                        for col_index in columns_to_keep_as_strings:
                            if values[col_index] is not None:
                                values[col_index] = str(values[col_index])

                        for i in range(11):
                            if i not in columns_to_keep_as_strings and values[i] is not None:
                                values[i] = int(values[i]) if values[i].isdigit() else values[i]

                        if str(values[0]).find('(') == 0:
                            values[0] = values[0].replace('(', '')

                        records_to_insert.append({
                            'index_id': int(values[0]),
                            'ipaddress': str(values[1]),
                            'DateVisited': str(values[2]),
                            'Neuronname': str(values[4]),
                            'neuron_id': int(values[10])
                        })

                        if len(records_to_insert) >= batch_size:
                            NMO_Log_Download.insert_many(records_to_insert)
                            records_to_insert = []

        if records_to_insert:
            NMO_Log_Download.insert_many(records_to_insert)

        return make_response(jsonify({'message': 'Data imported successfully'}), 200)


      except Exception as e:
        print('An error occurred:', str(e))
        return make_response(jsonify({'error': 'Data import failed'}), 500)



@app.route('/users/article_neuron_download', methods=['POST'])
def article_neuron_dowload():

    response = Response(status=200)

    NMO_Neuron_Article = db.NMO_Neuron_Article
    NMO_Log_Download = db.NMO_Log_Download
    NMO_Neuron_distinct_pmids = db.NMO_Neuron_distinct_pmids
    NMO_Neuron_distinct_pmids.drop()
    NMO_Neuron_distinct_downloads = db.NMO_Neuron_distinct_downloads
    NMO_Neuron_distinct_downloads.drop()

    db.create_collection('NMO_Neuron_distinct_pmids')
    # Use the aggregation framework to group by PMID and insert into the new collection
    pipeline1 = [
        {
            '$group': {
                '_id': '$PMID',
                'neuron_ids': {'$push': '$neuron_id'},
                'article_id': {'$addToSet': '$article_id'}
            }
        },
        {
            '$out': 'NMO_Neuron_distinct_pmids'
        }
    ]

    db.NMO_Neuron_Article.aggregate(pipeline1)

    db.create_collection('NMO_Neuron_distinct_downloads')
    # Use the aggregation framework to group by neuron_id and accumulate unique values
    pipeline2 = [
        {
            '$group': {
                '_id': '$neuron_id',
                'DateVisited': {'$push': '$DateVisited'},
                'Neuronname': {'$addToSet': '$Neuronname'},
                'count_DateVisited': {'$sum': 1}  # Count DateVisited
            }
        },
        {
            '$out': 'NMO_Neuron_distinct_downloads'
        }
    ]

    db.NMO_Log_Download.aggregate(pipeline2, allowDiskUse=True)

    return make_response(jsonify({'message': 'Collection created successfully'}), 200)


@app.route('/users/number_of_downloads_per_artcle', methods=['POST'])
def number_of_downloads_per_article():

    response = Response(status=200)

    NMO_Neuron_distinct_pmids = db.NMO_Neuron_distinct_pmids
    NMO_Neuron_distinct_downloads = db.NMO_Neuron_distinct_downloads
    NMO_Number_of_downloads_per_pmid = db.NMO_Number_of_downloads_per_pmid
    NMO_Number_of_downloads_per_pmid.drop()

    article_docs = NMO_Neuron_distinct_pmids.find()
    for doc1 in article_docs:
        neuron_downloads = []
        num_Neurons = 0
        neuron_id_total_downloads = 0
        avgDownloadsPerCell = 0
        for elem in doc1['neuron_ids']:
            res = NMO_Neuron_distinct_downloads.find_one({'_id' : elem})

            if res != None:
                num_Neurons += 1
                neuron_downloads_Obj = {}
                neuron_id_downloads = res['count_DateVisited']
                neuron_id_total_downloads += neuron_id_downloads
                neuron_downloads_Obj = {'neuron_id': elem, 'num_downloads': neuron_id_downloads}
                neuron_downloads.append(neuron_downloads_Obj)

        if len(neuron_downloads) != 0:
                  avgDownloadsPerCell = round(neuron_id_total_downloads / num_Neurons)
                  new_rec_id = NMO_Number_of_downloads_per_pmid.insert_one({
                                    'pmid': doc1['_id'],
                                    'num_downloads' : neuron_downloads,
                                    'neuron_total_downloads' : neuron_id_total_downloads,
                                    'avgDownloadsPerCell' : avgDownloadsPerCell
                                })


    return make_response(jsonify({'message': 'Collection created successfully'}), 200)


@app.route('/users/NMODescribing_All_CitedBy_NMO_With_Downloads', methods=['POST'])
def NMODescribing_All_CitedBy_NMO_With_Downloads():

    response = Response(status=200)

    NMODescribing_Overall_CitedByNMO = db.NMODescribing_Overall_CitedByNMO
    NMO_Number_of_downloads_per_pmid = db.NMO_Number_of_downloads_per_pmid
    NMODescribing_Overall_CitedByNMO_With_Downloads = db.NMODescribing_Overall_CitedByNMO_With_Downloads
    NMODescribing_Overall_CitedByNMO_With_Downloads.drop()

    docs1 = NMODescribing_Overall_CitedByNMO.find()
    for doc1 in docs1:
        new_rec_id = NMODescribing_Overall_CitedByNMO_With_Downloads.insert_one({
                               'pmid': doc1['pmid'],
                               'doi' : doc1['doi'],
                               'citedBy': doc1['citedBy'],
                               'title': doc1['title'],
                               'publishedDate': doc1['publishedDate'],
                               'total_NMO_Citations': doc1['total_NMO_Citations'],
                               'Number_of_Downloads': None,
                               'Avg_Downloads_Per_Cell': None
                               })

    for doc2 in NMO_Number_of_downloads_per_pmid.find():
        NMODescribing_Overall_CitedByNMO_With_Downloads.update_one({"pmid": str(doc2["pmid"])}, { "$set":
                                                                                                 {"Number_of_Downloads": doc2['neuron_total_downloads'],

                                                                                                  "Avg_Downloads_Per_Cell": doc2['avgDownloadsPerCell']} })


    return make_response(jsonify({'message': 'Creation created successfully'}), 200)


@app.route('/users/NMO_DescribingCited_With_Downloads', methods=['POST'])
def NMO_DescribingCited_With_Downloads():

    response = Response(status=200)

    NMO_DescribingCited = db.NMO_DescribingCited
    NMO_Number_of_downloads_per_pmid = db.NMO_Number_of_downloads_per_pmid
    NMO_DescribingCited_With_Downloads = db.NMO_DescribingCited_With_Downloads
    NMO_DescribingCited_With_Downloads.drop()

    docs1 = NMO_DescribingCited.find()
    for doc1 in docs1:
        new_rec_id = NMO_DescribingCited_With_Downloads.insert_one({
                                                'doi': doc1['doi'],
                                                'pmid': doc1['pmid'],
                                                'title': doc1['title'],
                                                'publishedDate': doc1['publishedDate'],
                                                'Number_of_Downloads': None,
                                                'Avg_Downloads_Per_Cell': None
                                                })

    for doc2 in NMO_Number_of_downloads_per_pmid.find():
        id = str(doc2['pmid']) + str(' *')
        NMO_DescribingCited_With_Downloads.update_one({"pmid": id}, { "$set":
                                                                    { "Number_of_Downloads": doc2['neuron_total_downloads'],
                                                                      "Avg_Downloads_Per_Cell": doc2['avgDownloadsPerCell']} })

    return make_response(jsonify({'message': 'Collection created successfully'}), 200)


if __name__ == '__main__':
    app.run(host='0.0.0.0')

